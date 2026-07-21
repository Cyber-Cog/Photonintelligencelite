"""Excel report generation (openpyxl — locked in docs/PRD.md §25).

Sheets: Overview, KPIs, Validation, one sheet per algorithm result (tables), Appendix
(mapping confidence + thresholds used). Mirrors the PDF's content so both formats never
present different numbers for the same job. See docs/PRD.md §7.11.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analytics.core.result import ResultStatus
from analytics.reports.report_data import ReportContext

_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=16, bold=True, color="0F172A")


def _write_table(ws, header: list[str], rows: list[list], start_row: int = 1) -> int:
    for col_idx, name in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24
    return start_row + len(rows) + 2


def build_excel(ctx: ReportContext) -> bytes:
    """Build the Excel workbook bytes for a job report (same numbers as the PDF)."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = ctx.report_title
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Generated {ctx.generated_at} · Job {ctx.job_id}"
    _write_table(
        ws,
        ["Field", "Value"],
        [
            ["Plant name", ctx.plant.plant_name],
            ["Timezone", ctx.plant.timezone],
            ["AC capacity (MW)", ctx.plant.ac_capacity_mw],
            ["DC capacity (MWp)", ctx.plant.dc_capacity_mwp],
            ["Module rating (Wp)", ctx.plant.module_rating_wp],
            ["Module technology", ctx.plant.module_technology],
            ["Bifacial", "Yes" if ctx.plant.bifacial else "No"],
            ["Tariff (INR/kWh)", ctx.plant.tariff_inr_per_kwh or "Not provided"],
            ["PR benchmark (%)", ctx.plant.pr_benchmark_pct or "Not provided"],
            ["External irradiance fallback used", "Yes" if ctx.external_irradiance_used else "No"],
        ],
        start_row=4,
    )

    ws_kpi = wb.create_sheet("KPIs")
    k = ctx.run.kpis
    _write_table(
        ws_kpi,
        ["KPI", "Value"],
        [
            ["Plant Availability (%)", k.get("plant_availability_pct")],
            ["Performance Ratio (%)", k.get("performance_ratio_pct")],
            ["Specific Yield (kWh/kWp)", k.get("specific_yield_kwh_per_kwp")],
            ["Estimated Energy Loss (kWh)", k.get("estimated_energy_loss_kwh")],
            ["Revenue Loss (INR)", k.get("revenue_loss_inr") if k.get("revenue_loss_available") else "Not available (no tariff supplied)"],
            ["Fault Count", k.get("fault_count", 0)],
            ["Total AC Energy (kWh)", k.get("total_ac_energy_kwh")],
        ],
    )

    ws_val = wb.create_sheet("Validation")
    v = ctx.validation
    ws_val["A1"] = f"{v.row_count} rows, {v.column_count} columns — {len(v.blockers)} blocker(s), {len(v.warnings)} warning(s)"
    ws_val["A2"] = f"Detected sampling interval: {ctx.interval_result.detected_interval_minutes:.1f} minutes"
    _write_table(
        ws_val,
        ["Code", "Severity", "Message", "Likely cause", "Affected rows"],
        [[i.code, i.severity, i.message, i.likely_cause, i.affected_rows] for i in v.issues],
        start_row=4,
    )

    for r in ctx.run.results:
        sheet_name = r.algorithm_id[:28]
        ws_r = wb.create_sheet(sheet_name)
        ws_r["A1"] = r.title
        ws_r["A1"].font = Font(bold=True, size=13)
        ws_r["A2"] = f"Status: {r.status.value} — {r.summary}"
        row_cursor = 4
        if r.status == ResultStatus.OK:
            metric_rows = [[k2.replace("_", " ").title(), v2] for k2, v2 in r.metrics.items()]
            row_cursor = _write_table(ws_r, ["Metric", "Value"], metric_rows, start_row=row_cursor)
            for table in r.tables:
                ws_r.cell(row=row_cursor, column=1, value=table.title).font = Font(bold=True)
                row_cursor += 1
                row_cursor = _write_table(ws_r, table.columns, table.rows, start_row=row_cursor)

    ws_app = wb.create_sheet("Appendix")
    ws_app["A1"] = "Column mapping confidence"
    ws_app["A1"].font = Font(bold=True)
    row_cursor = _write_table(ws_app, ["Source column", "Confidence"], [[c, f"{conf * 100:.0f}%"] for c, conf in ctx.mapping_confidence_by_column.items()], start_row=2)
    ws_app.cell(row=row_cursor, column=1, value="Non-default thresholds applied").font = Font(bold=True)
    row_cursor += 1
    threshold_rows = []
    for r in ctx.run.results:
        for tk, tv in r.thresholds_used.items():
            threshold_rows.append([r.algorithm_id, tk, tv])
    _write_table(ws_app, ["Algorithm", "Threshold", "Value"], threshold_rows, start_row=row_cursor)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_report(ctx: ReportContext) -> bytes:
    """Public alias used by the job worker (matches build_pdf_report naming)."""
    return build_excel(ctx)
