"""Plant architecture Excel template (download) and parser (upload).

Supported input shapes (all normalize to plant_config architecture JSON):

1. **Flat** (recommended for Indian sites) — one row per SMB/SCB:
   ``Inverter ID, SCB/SMB ID, Inverter kW, Strings per SCB`` (+ optional DC kWp / String ID)

2. **Embedded / companion** — same workbook as SCADA: sheet named Architecture /
   Plant / Master / Inverter List / Equipment, *or* SCADA columns Inverter ID + SCB ID
   (unique pairs derived when no dedicated sheet parses).

3. **Hierarchy** (Complete Analysis Pack / advanced) — Plant → Inverter → SCB → String:
   ``id, parent_id, device_type, ac_capacity_kw, dc_capacity_kwp, strings_per_scb, notes``

Plant AC/DC on hierarchy sheets use **kW / kWp**; PlantConfig stores **MW / MWp**
(÷ 1000). Inverter AC ratings land in ``equipment_ratings`` (kW).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Optional, Union

from openpyxl import Workbook, load_workbook

TEMPLATE_COLUMNS = (
    "inverter_id",
    "inverter_rated_kw",
    "scb_id",
    "strings_per_scb",
    "notes",
)

HIERARCHY_COLUMNS = (
    "id",
    "parent_id",
    "device_type",
    "ac_capacity_kw",
    "dc_capacity_kwp",
    "strings_per_scb",
    "notes",
)

TEMPLATE_FILENAME = "pic_lite_plant_architecture_template.xlsx"

_ARCH_SHEET_HINTS = (
    "architecture",
    "plant",
    "master",
    "inverter list",
    "inverter_list",
    "inv list",
    "equipment",
    "hierarchy",
    "plant layout",
    "layout",
    "smb list",
    "scb list",
    "equipment list",
    "plant master",
    "device list",
    "inverter master",
    "array",
    "topology",
)

_SKIP_SHEET_NAMES = {
    "instructions",
    "readme",
    "fault_checklist",
    "checklist",
    "cover",
    "index",
}

_COLUMN_SYNONYMS: dict[str, tuple[str, ...]] = {
    "inverter_id": (
        "inverter_id", "inverter", "inv", "inv_id", "inv_no", "inverter_no",
        "inverter_number", "inverter_name", "inverter_code", "inv_name",
        "inverterid", "pcs", "pcs_id", "ht_inverter",
    ),
    "scb_id": (
        "scb_id", "scb", "smb", "smb_id", "scb_no", "smb_no", "scb_number",
        "smb_number", "combiner", "combiner_id", "combiner_box", "combiner_no",
        "mppt", "mppt_id", "scbid", "smbid", "smb_name", "scb_name",
        "string_combiner", "ajb", "ajb_id",
    ),
    "string_id": (
        "string_id", "string", "str", "str_id", "string_no", "string_number",
        "string_name", "ct", "ct_id", "channel",
    ),
    "inverter_rated_kw": (
        "inverter_rated_kw", "inverter_rated_ac_kw", "rated_kw", "rating_kw",
        "inverter_kw", "inv_kw", "inv_rating", "inverter_rating",
        "inverter_rating_kw", "ac_rating_kw", "capacity_kw", "rated_ac_kw",
        "kw_rating", "inverter_capacity_kw", "pcs_kw",
    ),
    "strings_per_scb": (
        "strings_per_scb", "strings_per_smb", "no_of_strings", "number_of_strings",
        "num_strings", "string_count", "strings", "str_count", "strings_count",
        "no_strings",
    ),
    "dc_capacity_kwp": (
        "dc_capacity_kwp", "dc_capacity_kw", "dc_capacity", "dc_kwp", "dc_kw",
        "dc_rating_kwp", "dc_rating", "capacity_kwp", "kwp", "dc_cap_kwp",
    ),
    "id": ("id", "node_id", "equipment_id", "device_id", "asset_id", "tag"),
    "parent_id": ("parent_id", "parent", "parent_device_id", "parent_equipment_id"),
    "device_type": ("device_type", "type", "level", "equipment_type", "device_level", "node_type"),
    "ac_capacity_kw": ("ac_capacity_kw", "ac_capacity", "ac_kw", "ac_rating_kw", "plant_ac_kw"),
    "notes": ("notes", "remark", "remarks", "comment", "comments", "description"),
    "timestamp": ("timestamp", "time", "datetime", "date_time", "date_&_time", "date"),
}

_DEVICE_TYPE_ALIASES = {
    "plant": "plant",
    "site": "plant",
    "icr": "icr",
    "inverter control room": "icr",
    "pcs room": "icr",
    "inverter": "inverter",
    "inv": "inverter",
    "scb": "scb",
    "smb": "scb",
    "combiner": "scb",
    "mppt": "scb",
    "string": "string",
    "str": "string",
}


@dataclass
class ArchitectureParseResult:
    equipment_ratings: dict[str, float] = field(default_factory=dict)
    architecture: dict[str, dict] = field(default_factory=dict)
    inverters: list[dict] = field(default_factory=list)
    row_count: int = 0
    notes: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    plant_name: Optional[str] = None
    ac_capacity_mw: Optional[float] = None
    dc_capacity_mwp: Optional[float] = None
    inverter_capacity_kw: Optional[float] = None
    strings_per_scb: Optional[int] = None
    format: Optional[str] = None  # "hierarchy" | "flat" | "scada_embedded"
    source_sheet: Optional[str] = None

    @property
    def ok(self) -> bool:
        return bool(self.architecture) and not self.errors

    def to_plant_config_draft(self) -> dict:
        draft: dict = {
            "equipment_ratings": dict(self.equipment_ratings),
            "architecture": {
                scb_id: {
                    "inverter_id": entry["inverter_id"],
                    **(
                        {"strings_per_scb": entry["strings_per_scb"]}
                        if entry.get("strings_per_scb") is not None
                        else {}
                    ),
                    **(
                        {"modules_per_string": entry["modules_per_string"]}
                        if entry.get("modules_per_string") is not None
                        else {}
                    ),
                    **(
                        {"dc_capacity_kwp": entry["dc_capacity_kwp"]}
                        if entry.get("dc_capacity_kwp") is not None
                        else {}
                    ),
                }
                for scb_id, entry in self.architecture.items()
            },
            "architecture_imported": True,
            "architecture_format": self.format,
        }
        if self.plant_name:
            draft["plant_name"] = self.plant_name
        if self.ac_capacity_mw is not None and self.ac_capacity_mw > 0:
            draft["ac_capacity_mw"] = self.ac_capacity_mw
        if self.dc_capacity_mwp is not None and self.dc_capacity_mwp > 0:
            draft["dc_capacity_mwp"] = self.dc_capacity_mwp
        if self.inverter_capacity_kw is not None and self.inverter_capacity_kw > 0:
            draft["inverter_capacity_kw"] = self.inverter_capacity_kw
        elif self.equipment_ratings:
            draft["inverter_capacity_kw"] = max(self.equipment_ratings.values())
        if self.strings_per_scb is not None:
            draft["strings_per_scb"] = self.strings_per_scb
        return draft


def build_template_bytes(
    *,
    example_inverters: int = 2,
    scbs_per_inverter: int = 16,
    strings_per_scb: int = 24,
    default_rated_kw: float = 1500.0,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "architecture"
    ws.append(list(TEMPLATE_COLUMNS))

    for inv_i in range(1, example_inverters + 1):
        inv_id = f"INV-{inv_i:02d}"
        for scb_i in range(1, scbs_per_inverter + 1):
            scb_id = f"{inv_id}-SCB-{scb_i:02d}"
            notes = "Example — replace with your plant IDs" if inv_i == 1 and scb_i == 1 else ""
            ws.append([inv_id, default_rated_kw, scb_id, strings_per_scb, notes])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    instr = wb.create_sheet("instructions", 0)
    instr["A1"] = "PIC Lite — Plant architecture (flat table — recommended for Indian sites)"
    instr["A3"] = "How to use"
    instr["A4"] = "1. Keep the header row on the 'architecture' sheet (synonyms OK: INV, SCB/SMB, Rating kW, No of Strings)."
    instr["A5"] = "2. One row per SMB/SCB. Repeat inverter_id and inverter_rated_kw on every SCB row for that inverter."
    instr["A6"] = "3. inverter_id / scb_id should match Device IDs in your SCADA export."
    instr["A7"] = "4. strings_per_scb = number of strings feeding that SMB (needed for current-clipping / disconnected strings)."
    instr["A8"] = "5. Optional: add dc_capacity_kwp or String ID columns — String IDs count strings when strings_per_scb is blank."
    instr["A9"] = "6. You can put this sheet in the SAME workbook as SCADA (sheet names: Architecture, Plant, Master, Inverter List…)."
    instr["A10"] = "7. Delete the example rows, paste your plant, save as .xlsx, upload with SCADA or in Setup."
    instr["A12"] = "Advanced (hierarchy pack): id, parent_id, device_type, ac_capacity_kw, dc_capacity_kwp, strings_per_scb."
    instr["A13"] = "See docs/ARCHITECTURE_INGEST.md for detection rules and Method A / B / C."
    instr.column_dimensions["A"].width = 110

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _float_or_none(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        v = float(value)
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


def _int_or_none(value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        v = int(float(value))
        return v if v >= 1 else None
    except (TypeError, ValueError):
        return None


def _normalize_header(value) -> str:
    s = (_cell(value) or "").lower()
    for ch in (" ", "-", "/", ".", "(", ")", "[", "]"):
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def _normalize_device_type(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    key = value.strip().lower().replace(" ", "_").replace("-", "_")
    return _DEVICE_TYPE_ALIASES.get(key)


def _kw_to_mw(kw: Optional[float]) -> Optional[float]:
    if kw is None:
        return None
    return round(kw / 1000.0, 6)


def _sheet_name_score(name: str) -> int:
    lower = (name or "").strip().lower()
    if lower in _SKIP_SHEET_NAMES:
        return -100
    if lower == "architecture":
        return 100
    if lower in {"scada", "telemetry", "data", "raw"}:
        return -20
    for hint in _ARCH_SHEET_HINTS:
        if lower == hint or hint in lower:
            return 80
    return 0


def _map_headers_to_canonical(raw_headers: list) -> dict[str, int]:
    normalized = [_normalize_header(h) for h in raw_headers]
    synonym_to_canon: dict[str, str] = {}
    for canon, syns in _COLUMN_SYNONYMS.items():
        for syn in syns:
            synonym_to_canon.setdefault(syn, canon)

    col_idx: dict[str, int] = {}
    for i, norm in enumerate(normalized):
        if not norm:
            continue
        if norm in _COLUMN_SYNONYMS and norm not in col_idx:
            col_idx[norm] = i
            continue
        canon = synonym_to_canon.get(norm)
        if canon and canon not in col_idx:
            col_idx[canon] = i
    return col_idx


def _detect_layout(col_idx: dict[str, int]) -> Optional[str]:
    if "id" in col_idx and ("device_type" in col_idx or "parent_id" in col_idx):
        return "hierarchy"
    if "inverter_id" in col_idx and "scb_id" in col_idx:
        return "flat"
    return None


def _looks_like_scada_sheet(col_idx: dict[str, int]) -> bool:
    return "timestamp" in col_idx


def _col(raw: tuple, col_idx: dict[str, int], name: str):
    i = col_idx.get(name)
    if i is None or i >= len(raw):
        return None
    return raw[i]


def _read_sheet_header(ws) -> tuple[Optional[list], dict[str, int]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        return None, {}
    header = list(header_raw)
    return header, _map_headers_to_canonical(header)


def _open_workbook(source: Union[Path, BinaryIO, bytes]):
    if isinstance(source, bytes):
        return load_workbook(BytesIO(source), read_only=True, data_only=True)
    if isinstance(source, Path):
        return load_workbook(source, read_only=True, data_only=True)
    return load_workbook(source, read_only=True, data_only=True)


def _parse_flat(rows, col_idx: dict[str, int], result: ArchitectureParseResult) -> ArchitectureParseResult:
    inv_order: list[str] = []
    inv_map: dict[str, dict] = {}
    string_ids_by_scb: dict[str, set[str]] = {}

    for raw in rows:
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        inv_id = _cell(_col(raw, col_idx, "inverter_id"))
        scb_id = _cell(_col(raw, col_idx, "scb_id"))
        if not inv_id or not scb_id:
            continue

        rated = _float_or_none(_col(raw, col_idx, "inverter_rated_kw"))
        if rated is None:
            rated = _float_or_none(_col(raw, col_idx, "ac_capacity_kw"))
        strings = _int_or_none(_col(raw, col_idx, "strings_per_scb"))
        dc_kwp = _float_or_none(_col(raw, col_idx, "dc_capacity_kwp"))
        string_id = _cell(_col(raw, col_idx, "string_id"))

        result.row_count += 1
        if inv_id not in inv_map:
            inv_order.append(inv_id)
            inv_map[inv_id] = {"inverter_id": inv_id, "rated_kw": rated, "scbs": {}}
        elif rated is not None and inv_map[inv_id]["rated_kw"] is None:
            inv_map[inv_id]["rated_kw"] = rated

        if scb_id not in inv_map[inv_id]["scbs"]:
            inv_map[inv_id]["scbs"][scb_id] = {
                "scb_id": scb_id,
                "strings_per_scb": strings,
                "strings_detected": False,
            }
        elif strings is not None and inv_map[inv_id]["scbs"][scb_id].get("strings_per_scb") is None:
            inv_map[inv_id]["scbs"][scb_id]["strings_per_scb"] = strings

        entry: dict = {"inverter_id": inv_id}
        existing = result.architecture.get(scb_id)
        if existing and existing.get("strings_per_scb") is not None and strings is None:
            strings = existing["strings_per_scb"]
        if strings is not None:
            entry["strings_per_scb"] = strings
        if dc_kwp is not None:
            entry["dc_capacity_kwp"] = dc_kwp
        elif existing and existing.get("dc_capacity_kwp") is not None:
            entry["dc_capacity_kwp"] = existing["dc_capacity_kwp"]
        result.architecture[scb_id] = entry
        if rated is not None:
            result.equipment_ratings[inv_id] = rated
        if string_id:
            string_ids_by_scb.setdefault(scb_id, set()).add(string_id)
            inv_map[inv_id]["scbs"][scb_id]["strings_detected"] = True

    for scb_id, ids in string_ids_by_scb.items():
        if scb_id in result.architecture and result.architecture[scb_id].get("strings_per_scb") is None:
            result.architecture[scb_id]["strings_per_scb"] = len(ids)
            for inv in inv_map.values():
                if scb_id in inv["scbs"] and inv["scbs"][scb_id].get("strings_per_scb") is None:
                    inv["scbs"][scb_id]["strings_per_scb"] = len(ids)

    if not result.architecture:
        result.errors.append(
            "No valid architecture rows found. Each row needs Inverter ID and SCB/SMB ID "
            "(or synonyms like INV / SMB)."
        )
        return result

    for inv_id in inv_order:
        node = inv_map[inv_id]
        result.inverters.append(
            {
                "inverter_id": inv_id,
                "rated_kw": node["rated_kw"],
                "scbs": list(node["scbs"].values()),
            }
        )

    if result.equipment_ratings:
        result.inverter_capacity_kw = max(result.equipment_ratings.values())

    string_counts = [
        e["strings_per_scb"]
        for e in result.architecture.values()
        if e.get("strings_per_scb") is not None
    ]
    if string_counts:
        result.strings_per_scb = max(set(string_counts), key=string_counts.count)

    scb_dc = [
        e["dc_capacity_kwp"]
        for e in result.architecture.values()
        if e.get("dc_capacity_kwp") is not None
    ]
    if scb_dc and result.dc_capacity_mwp is None:
        result.dc_capacity_mwp = _kw_to_mw(sum(scb_dc))

    n_inv = len(result.inverters)
    n_scb = len(result.architecture)
    n_rated = len(result.equipment_ratings)
    result.notes.append(
        f"Loaded {n_inv} inverter(s), {n_scb} SMB/SCB(s) from {result.row_count} Excel row(s)"
        + (f"; ratings for {n_rated} inverter(s)." if n_rated else ".")
    )
    return result


def _parse_hierarchy(rows, col_idx: dict[str, int], result: ArchitectureParseResult) -> ArchitectureParseResult:
    inv_order: list[str] = []
    inv_map: dict[str, dict] = {}
    scb_parent_inv: dict[str, str] = {}
    string_counts: list[int] = []

    for raw in rows:
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        node_id = _cell(_col(raw, col_idx, "id"))
        if not node_id:
            continue
        parent_id = _cell(_col(raw, col_idx, "parent_id"))
        dtype = _normalize_device_type(_cell(_col(raw, col_idx, "device_type")))
        ac_kw = _float_or_none(_col(raw, col_idx, "ac_capacity_kw"))
        dc_kwp = _float_or_none(_col(raw, col_idx, "dc_capacity_kwp"))
        strings = _int_or_none(_col(raw, col_idx, "strings_per_scb"))
        result.row_count += 1

        if dtype is None:
            lower = node_id.lower()
            if "str" in lower.split("-")[-1] or lower.endswith("-str") or "-str-" in lower:
                dtype = "string"
            elif "scb" in lower or "smb" in lower:
                dtype = "scb"
            elif lower.startswith("inv") or "-inv-" in lower:
                dtype = "inverter"
            elif lower.startswith("plant") or lower in {"site", "plant"}:
                dtype = "plant"
            else:
                result.notes.append(f"Skipped row id={node_id!r}: unknown device_type.")
                continue

        if dtype == "plant":
            note = _cell(_col(raw, col_idx, "notes"))
            if not result.plant_name:
                result.plant_name = note or node_id
            if ac_kw is not None:
                result.ac_capacity_mw = _kw_to_mw(ac_kw)
            if dc_kwp is not None:
                result.dc_capacity_mwp = _kw_to_mw(dc_kwp)
            continue

        if dtype == "inverter":
            if node_id not in inv_map:
                inv_order.append(node_id)
                inv_map[node_id] = {"inverter_id": node_id, "rated_kw": ac_kw, "scbs": {}, "dc_kwp": dc_kwp}
            else:
                if ac_kw is not None and inv_map[node_id]["rated_kw"] is None:
                    inv_map[node_id]["rated_kw"] = ac_kw
                if dc_kwp is not None:
                    inv_map[node_id]["dc_kwp"] = dc_kwp
            if ac_kw is not None:
                result.equipment_ratings[node_id] = ac_kw
            continue

        if dtype == "scb":
            inv_id = parent_id
            if inv_id and inv_id in inv_map:
                pass
            elif inv_id and inv_id not in inv_map:
                inv_order.append(inv_id)
                inv_map[inv_id] = {"inverter_id": inv_id, "rated_kw": None, "scbs": {}}
            else:
                result.notes.append(f"SCB {node_id!r} missing parent_id (inverter); skipped.")
                continue

            scb_parent_inv[node_id] = inv_id  # type: ignore[assignment]
            inv_map[inv_id]["scbs"][node_id] = {  # type: ignore[index]
                "scb_id": node_id,
                "strings_per_scb": strings,
                "strings_detected": False,
                "dc_capacity_kwp": dc_kwp,
            }
            entry: dict = {"inverter_id": inv_id}  # type: ignore[assignment]
            if strings is not None:
                entry["strings_per_scb"] = strings
                string_counts.append(strings)
            if dc_kwp is not None:
                entry["dc_capacity_kwp"] = dc_kwp
            if ac_kw is not None:
                entry["ac_capacity_kw"] = ac_kw
            result.architecture[node_id] = entry
            continue

        if dtype == "string":
            scb_id = parent_id
            if not scb_id:
                continue
            inv_id = scb_parent_inv.get(scb_id)
            if inv_id is None:
                continue
            scb_node = inv_map.get(inv_id, {}).get("scbs", {}).get(scb_id)
            if scb_node is not None:
                scb_node["strings_detected"] = True
                counted = scb_node.setdefault("_string_ids", set())
                counted.add(node_id)
            continue

    for inv_id, node in inv_map.items():
        for scb_id, scb in node["scbs"].items():
            counted = scb.pop("_string_ids", None)
            if scb.get("strings_per_scb") is None and counted:
                scb["strings_per_scb"] = len(counted)
                if scb_id in result.architecture:
                    result.architecture[scb_id]["strings_per_scb"] = len(counted)
                    string_counts.append(len(counted))

    if not result.architecture:
        if inv_map:
            result.errors.append(
                "Hierarchy sheet has inverters but no SCB rows. Add device_type=scb rows "
                "with parent_id pointing at each inverter."
            )
        else:
            result.errors.append(
                "No valid hierarchy rows found. Expected columns: "
                + ", ".join(HIERARCHY_COLUMNS)
                + "."
            )
        return result

    for inv_id in inv_order:
        node = inv_map[inv_id]
        result.inverters.append(
            {
                "inverter_id": inv_id,
                "rated_kw": node["rated_kw"],
                "scbs": [
                    {
                        "scb_id": s["scb_id"],
                        "strings_per_scb": s.get("strings_per_scb"),
                        "strings_detected": bool(s.get("strings_detected")),
                    }
                    for s in node["scbs"].values()
                ],
            }
        )

    if result.equipment_ratings and result.inverter_capacity_kw is None:
        result.inverter_capacity_kw = max(result.equipment_ratings.values())

    if result.ac_capacity_mw is None and result.equipment_ratings:
        result.ac_capacity_mw = _kw_to_mw(sum(result.equipment_ratings.values()))
    if result.dc_capacity_mwp is None:
        inv_dc = [
            node.get("dc_kwp")
            for node in inv_map.values()
            if node.get("dc_kwp") is not None
        ]
        if inv_dc:
            result.dc_capacity_mwp = _kw_to_mw(sum(inv_dc))  # type: ignore[arg-type]
        else:
            scb_dc = [
                e["dc_capacity_kwp"]
                for e in result.architecture.values()
                if e.get("dc_capacity_kwp") is not None
            ]
            if scb_dc:
                result.dc_capacity_mwp = _kw_to_mw(sum(scb_dc))

    if string_counts:
        result.strings_per_scb = max(set(string_counts), key=string_counts.count)

    n_inv = len(result.inverters)
    n_scb = len(result.architecture)
    n_rated = len(result.equipment_ratings)
    cap_bits = []
    if result.ac_capacity_mw is not None:
        cap_bits.append(f"plant AC {result.ac_capacity_mw:g} MW")
    if result.dc_capacity_mwp is not None:
        cap_bits.append(f"DC {result.dc_capacity_mwp:g} MWp")
    result.notes.append(
        f"Loaded hierarchy: {n_inv} inverter(s), {n_scb} SMB/SCB(s) from {result.row_count} row(s)"
        + (f"; ratings for {n_rated}." if n_rated else ".")
        + ((" (" + ", ".join(cap_bits) + ")") if cap_bits else "")
    )
    return result


def parse_architecture_excel(source: Union[Path, BinaryIO, bytes]) -> ArchitectureParseResult:
    result = ArchitectureParseResult()
    raw_bytes: Optional[bytes] = None
    try:
        if isinstance(source, bytes):
            raw_bytes = source
        elif not isinstance(source, Path) and hasattr(source, "read"):
            raw_bytes = source.read()
        wb = _open_workbook(raw_bytes if raw_bytes is not None else source)
    except Exception as exc:  # noqa: BLE001
        result.errors.append(f"Could not read Excel file: {exc}")
        return result

    best: Optional[tuple[int, str, bool]] = None
    try:
        for name in wb.sheetnames:
            name_score = _sheet_name_score(name)
            if name_score <= -100:
                continue
            ws = wb[name]
            header, col_idx = _read_sheet_header(ws)
            if header is None or not col_idx:
                continue
            layout = _detect_layout(col_idx)
            if layout is None:
                continue
            is_scada = _looks_like_scada_sheet(col_idx)
            score = name_score
            if is_scada:
                score -= 40
            if layout == "hierarchy":
                score += 5
            if layout == "flat" and not is_scada:
                score += 10
            if best is None or score > best[0]:
                best = (score, name, is_scada)
    finally:
        wb.close()

    if best is None:
        result.errors.append(
            "No architecture sheet found. Add a flat table (Inverter ID, SCB/SMB ID, …) "
            "on a sheet named Architecture / Plant / Master, or include hierarchy columns, "
            "or put Inverter ID + SCB ID on the SCADA sheet."
        )
        return result

    _, sheet_name, is_scada = best

    try:
        if raw_bytes is not None:
            wb2 = _open_workbook(raw_bytes)
        elif isinstance(source, Path):
            wb2 = _open_workbook(source)
        elif isinstance(source, bytes):
            wb2 = _open_workbook(source)
        else:
            result.errors.append("Could not re-open workbook to parse architecture sheet.")
            return result
    except Exception as exc:  # noqa: BLE001
        result.errors.append(f"Could not read Excel file: {exc}")
        return result

    try:
        ws = wb2[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows_iter)
        except StopIteration:
            result.errors.append("Architecture sheet is empty.")
            return result
        col_idx = _map_headers_to_canonical(list(header_raw))
        layout = _detect_layout(col_idx)
        if layout is None:
            result.errors.append(
                "Unrecognized architecture columns. Expected flat "
                "(Inverter ID + SCB/SMB ID …) or hierarchy "
                f"({', '.join(HIERARCHY_COLUMNS)})."
            )
            return result
        result = ArchitectureParseResult(source_sheet=sheet_name)
        result.format = "scada_embedded" if (is_scada and layout == "flat") else layout
        if layout == "hierarchy":
            parsed = _parse_hierarchy(rows_iter, col_idx, result)
        else:
            parsed = _parse_flat(rows_iter, col_idx, result)
        if parsed.ok and parsed.source_sheet and parsed.notes:
            parsed.notes[0] = parsed.notes[0].rstrip(".") + f" from sheet '{parsed.source_sheet}'."
        return parsed
    finally:
        wb2.close()


def workbook_has_architecture_sheet(source: Union[Path, BinaryIO, bytes]) -> bool:
    try:
        wb = _open_workbook(source)
    except Exception:  # noqa: BLE001
        return False
    try:
        for name in wb.sheetnames:
            if _sheet_name_score(name) >= 80:
                return True
            if name.strip().lower() in _SKIP_SHEET_NAMES:
                continue
            ws = wb[name]
            header, col_idx = _read_sheet_header(ws)
            if header is None:
                continue
            if _detect_layout(col_idx) is not None:
                return True
        return False
    finally:
        wb.close()


def try_parse_architecture_from_pack(source: Union[Path, BinaryIO, bytes]) -> Optional[ArchitectureParseResult]:
    parsed = parse_architecture_excel(source)
    if parsed.ok:
        return parsed
    return None


def apply_smb_pattern(
    inverter_ids: list[str],
    *,
    smbs_per_inverter: int,
    strings_per_smb: int,
    rated_kw: Optional[float] = None,
    existing: Optional[list[dict]] = None,
) -> list[dict]:
    if smbs_per_inverter < 1 or strings_per_smb < 1:
        raise ValueError("smbs_per_inverter and strings_per_smb must be >= 1")

    selected = {i.strip() for i in inverter_ids if i and i.strip()}
    preserved: list[dict] = []
    if existing:
        for inv in existing:
            if inv.get("inverter_id") not in selected:
                preserved.append(inv)

    generated: list[dict] = []
    for inv_id in sorted(selected):
        scbs = []
        for i in range(1, smbs_per_inverter + 1):
            scbs.append(
                {
                    "scb_id": f"{inv_id}-SCB-{i:02d}",
                    "strings_per_scb": strings_per_smb,
                    "strings_detected": False,
                }
            )
        generated.append({"inverter_id": inv_id, "rated_kw": rated_kw, "scbs": scbs})

    return preserved + generated


def inverters_from_plant_architecture(
    architecture: dict[str, dict],
    equipment_ratings: Optional[dict[str, float]] = None,
) -> list[dict]:
    ratings = equipment_ratings or {}
    inv_order: list[str] = []
    inv_map: dict[str, dict] = {}
    for scb_id, entry in architecture.items():
        inv_id = (entry or {}).get("inverter_id")
        if not inv_id:
            continue
        if inv_id not in inv_map:
            inv_order.append(inv_id)
            inv_map[inv_id] = {
                "inverter_id": inv_id,
                "rated_kw": ratings.get(inv_id),
                "scbs": [],
            }
        inv_map[inv_id]["scbs"].append(
            {
                "scb_id": scb_id,
                "strings_per_scb": (entry or {}).get("strings_per_scb"),
                "strings_detected": False,
            }
        )
    return [inv_map[i] for i in inv_order]
