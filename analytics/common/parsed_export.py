"""Export a job's parsed SCADA as Complete Analysis Pack–compatible Excel.

Round-trips through normal upload: sheet ``scada`` uses official SCADA_COLUMNS;
optional ``architecture`` (hierarchy) and ``column_mapping`` companion sheets.
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Sequence

from openpyxl import Workbook

from analytics.common.architecture_excel import HIERARCHY_COLUMNS
from analytics.common.aliasing import MEDIUM_CONFIDENCE, score_columns
from analytics.common.complete_analysis_pack import (
    OFFICIAL_COLUMN_TO_CANONICAL,
    SCADA_COLUMNS,
    _normalize_header_key,
)

# Canonical field names used in mapping_json / pipeline (timestamp, not timestamp_utc).
_CANONICAL_TO_OFFICIAL: dict[str, str] = {
    field: header for header, field in OFFICIAL_COLUMN_TO_CANONICAL.items()
}
# Post-standardize parquet uses timestamp_utc; map back to pack Timestamp header.
_CANONICAL_TO_OFFICIAL["timestamp_utc"] = "Timestamp"

# OEM exports often label the equipment column as Inverter / SMB / String — still one tidy ID column.
_DEVICE_ID_FIELD_ALIASES: tuple[str, ...] = ("device_id", "inverter_id", "scb_id", "string_id")

_DEFAULT_MAX_ROWS = 200_000


def _strip_bom(value: str) -> str:
    return (value or "").lstrip("\ufeff").strip()


def _header_name_lookup(raw_headers: Sequence[str]) -> dict[str, str]:
    """Normalized header key → actual CSV column name (first occurrence wins)."""
    out: dict[str, str] = {}
    for raw in raw_headers:
        actual = _strip_bom(str(raw))
        key = _normalize_header_key(actual)
        if key and key not in out:
            out[key] = actual
    return out


def _resolve_header_name(name: str, header_set: set[str], lookup: dict[str, str]) -> str | None:
    """Match a mapping or official name to the CSV column (exact, then normalized)."""
    if name in header_set:
        return name
    key = _normalize_header_key(name)
    if key in lookup:
        return lookup[key]
    return None


def _infer_column_to_canonical(raw_headers: Sequence[str]) -> dict[str, str]:
    """Best-effort mapping from CSV headers when job.mapping_json is empty or stale."""
    out: dict[str, str] = {}
    for candidate in score_columns([str(h) for h in raw_headers]):
        if (
            candidate.canonical_field
            and candidate.confidence >= MEDIUM_CONFIDENCE
            and candidate.column_name not in out
        ):
            out[candidate.column_name] = candidate.canonical_field
    return out


def _sanitize_filename_part(value: str, *, max_len: int = 40) -> str:
    cleaned = re.sub(r"[^\w.\-]+", "_", (value or "").strip(), flags=re.UNICODE)
    cleaned = cleaned.strip("._-") or "job"
    return cleaned[:max_len]


def parsed_export_filename(job_id: str, plant_name: str | None = None) -> str:
    """Filename for Content-Disposition / browser download."""
    short = (job_id or "job")[:8]
    if plant_name and plant_name.strip():
        return f"pic_lite_parsed_{_sanitize_filename_part(plant_name)}_{short}.xlsx"
    return f"pic_lite_parsed_{short}.xlsx"


def plant_name_from_config(plant_config_json: Mapping[str, Any] | None) -> Optional[str]:
    if not plant_config_json:
        return None
    plant = (
        plant_config_json.get("plant")
        if isinstance(plant_config_json.get("plant"), dict)
        else plant_config_json
    )
    if not isinstance(plant, dict):
        return None
    name = plant.get("plant_name")
    if name and str(name).strip():
        return str(name).strip()
    return None


def source_columns_for_official(
    column_to_canonical: Mapping[str, str] | None,
    timestamp_column: str | None,
    raw_headers: Sequence[str],
) -> dict[str, str | None]:
    """Map each official SCADA header → source CSV column name (or None)."""
    cleaned_headers = [_strip_bom(str(h)) for h in raw_headers]
    header_set = set(cleaned_headers)
    lookup = _header_name_lookup(cleaned_headers)
    c2c = {str(k): str(v) for k, v in (column_to_canonical or {}).items() if v and v != "ignore"}

    by_field: dict[str, str] = {}
    for src, field in c2c.items():
        actual = _resolve_header_name(src, header_set, lookup)
        if actual and field not in by_field:
            by_field[field] = actual

    if timestamp_column:
        ts_actual = _resolve_header_name(timestamp_column, header_set, lookup)
        if ts_actual:
            by_field.setdefault("timestamp", ts_actual)

    for alt in _DEVICE_ID_FIELD_ALIASES:
        if alt in by_field:
            by_field.setdefault("device_id", by_field[alt])
            break

    out: dict[str, str | None] = {}
    for official in SCADA_COLUMNS:
        field = OFFICIAL_COLUMN_TO_CANONICAL[official]
        src = by_field.get(field)
        if src is None and field == "device_id":
            for alt in _DEVICE_ID_FIELD_ALIASES[1:]:
                src = by_field.get(alt)
                if src:
                    break
        if src is None:
            src = _resolve_header_name(official, header_set, lookup)
        out[official] = src
    return out


def resolve_source_columns_for_official(
    column_to_canonical: Mapping[str, str] | None,
    timestamp_column: str | None,
    raw_headers: Sequence[str],
) -> dict[str, str | None]:
    """Resolve CSV columns for export; infer from headers when mapping is missing or stale."""
    cleaned_headers = [_strip_bom(str(h)) for h in raw_headers]
    resolved = source_columns_for_official(column_to_canonical, timestamp_column, cleaned_headers)
    if any(resolved.values()):
        return resolved

    inferred = _infer_column_to_canonical(cleaned_headers)
    ts_col = timestamp_column
    if not ts_col:
        for src, field in inferred.items():
            if field == "timestamp":
                ts_col = src
                break
    inferred_resolved = source_columns_for_official(inferred, ts_col, cleaned_headers)
    if any(inferred_resolved.values()):
        return inferred_resolved

    return resolved


def read_raw_csv_headers(raw_csv: Path) -> list[str]:
    """Read the first CSV row as column names (BOM-safe)."""
    with raw_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            raw_headers = next(reader)
        except StopIteration:
            return []
    return [_strip_bom(str(h)) for h in raw_headers]


def remap_csv_to_scada_rows(
    raw_csv: Path,
    source_for_official: Mapping[str, str | None],
    max_rows: int = _DEFAULT_MAX_ROWS,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    with raw_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for i, rec in enumerate(reader):
            if i >= max_rows:
                break
            out: list[Any] = []
            for official in SCADA_COLUMNS:
                src = source_for_official.get(official)
                if src is None:
                    out.append("")
                else:
                    val = rec.get(src, "")
                    out.append("" if val is None else val)
            rows.append(out)
    return rows


def canonical_frame_to_scada_rows(
    records: Iterable[Mapping[str, Any]],
    max_rows: int = _DEFAULT_MAX_ROWS,
) -> list[list[Any]]:
    """Convert canonical-field records to pack-header rows."""
    rows: list[list[Any]] = []
    for i, rec in enumerate(records):
        if i >= max_rows:
            break
        out: list[Any] = []
        for official in SCADA_COLUMNS:
            field = OFFICIAL_COLUMN_TO_CANONICAL[official]
            if field == "timestamp":
                val = rec.get("timestamp_utc", rec.get("timestamp", ""))
            elif field == "device_id":
                val = (
                    rec.get("device_id")
                    or rec.get("string_id")
                    or rec.get("scb_id")
                    or rec.get("inverter_id")
                    or ""
                )
            else:
                val = rec.get(field, "")
            if val is None:
                out.append("")
            elif isinstance(val, float) and val != val:  # NaN
                out.append("")
            else:
                out.append(val)
        rows.append(out)
    return rows


def architecture_hierarchy_rows(plant: Mapping[str, Any]) -> list[list[Any]]:
    """Rebuild hierarchy rows from plant_config plant dict."""
    architecture = plant.get("architecture") or {}
    if not isinstance(architecture, dict) or not architecture:
        return []

    ratings = plant.get("equipment_ratings") or {}
    plant_name = plant.get("plant_name") or "Plant"
    plant_id = "PLANT"

    ac_mw = plant.get("ac_capacity_mw")
    dc_mwp = plant.get("dc_capacity_mwp")
    ac_kw = float(ac_mw) * 1000.0 if ac_mw is not None else ""
    dc_kwp = float(dc_mwp) * 1000.0 if dc_mwp is not None else ""

    rows: list[list[Any]] = [
        [plant_id, "", "plant", ac_kw, dc_kwp, "", plant_name],
    ]

    inv_order: list[str] = []
    inv_seen: set[str] = set()
    scb_by_inv: dict[str, list[tuple[str, dict]]] = {}
    for scb_id, entry in architecture.items():
        if not isinstance(entry, dict):
            continue
        inv_id = entry.get("inverter_id")
        if not inv_id:
            continue
        inv_id = str(inv_id)
        if inv_id not in inv_seen:
            inv_seen.add(inv_id)
            inv_order.append(inv_id)
            scb_by_inv[inv_id] = []
        scb_by_inv[inv_id].append((str(scb_id), entry))

    default_rated = plant.get("inverter_capacity_kw")
    for inv_id in inv_order:
        rated = ratings.get(inv_id, default_rated)
        inv_dc: Any = ""
        scb_dcs = [
            float(e.get("dc_capacity_kwp"))
            for _, e in scb_by_inv.get(inv_id, [])
            if e.get("dc_capacity_kwp") is not None
        ]
        if scb_dcs:
            inv_dc = round(sum(scb_dcs), 6)
        rows.append([inv_id, plant_id, "inverter", rated if rated is not None else "", inv_dc, "", ""])

        for scb_id, entry in scb_by_inv.get(inv_id, []):
            strings = entry.get("strings_per_scb")
            scb_dc = entry.get("dc_capacity_kwp")
            scb_ac = entry.get("ac_capacity_kw")
            notes = "spare" if entry.get("spare_flag") else ""
            rows.append(
                [
                    scb_id,
                    inv_id,
                    "scb",
                    scb_ac if scb_ac is not None else "",
                    scb_dc if scb_dc is not None else "",
                    strings if strings is not None else "",
                    notes,
                ]
            )
            string_ids = entry.get("string_ids") or entry.get("strings") or []
            if isinstance(string_ids, list):
                for sid in string_ids:
                    rows.append([str(sid), scb_id, "string", "", "", "", ""])

    return rows


def _append_sheet_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18


def build_parsed_workbook_bytes(
    *,
    scada_rows: list[list[Any]],
    column_to_canonical: Mapping[str, str] | None = None,
    timestamp_column: str | None = None,
    plant: Mapping[str, Any] | None = None,
) -> bytes:
    """Assemble workbook from already-prepared scada rows + optional companions."""
    wb = Workbook()
    ws_scada = wb.active
    ws_scada.title = "scada"
    _append_sheet_rows(ws_scada, list(SCADA_COLUMNS), scada_rows)

    c2c = dict(column_to_canonical or {})
    if c2c or timestamp_column:
        ws_map = wb.create_sheet("column_mapping")
        ws_map.append(["Source column", "Canonical field", "Official pack header"])
        seen: set[str] = set()
        items = list(c2c.items())
        if timestamp_column and timestamp_column not in c2c:
            items.insert(0, (timestamp_column, "timestamp"))
        for src, field in items:
            if not field or field == "ignore":
                continue
            key = f"{src}|{field}"
            if key in seen:
                continue
            seen.add(key)
            official = _CANONICAL_TO_OFFICIAL.get(str(field), "")
            ws_map.append([src, field, official])
        for col in ws_map.columns:
            ws_map.column_dimensions[col[0].column_letter].width = 28

    if isinstance(plant, dict):
        arch_rows = architecture_hierarchy_rows(plant)
        if arch_rows:
            ws_arch = wb.create_sheet("architecture")
            _append_sheet_rows(ws_arch, list(HIERARCHY_COLUMNS), arch_rows)
            ws_arch.column_dimensions["G"].width = 36

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_parsed_excel_bytes(
    *,
    raw_csv: Path | None,
    mapping_json: Mapping[str, Any] | None = None,
    plant_config_json: Mapping[str, Any] | None = None,
    scada_rows_override: list[list[Any]] | None = None,
    max_rows: int = _DEFAULT_MAX_ROWS,
) -> bytes:
    """Build workbook from job raw CSV (preferred) or prebuilt scada rows.

    Sheets:
      - scada: official SCADA_COLUMNS (analysis-ready tidy long)
      - column_mapping: source → canonical → official (when mapping present)
      - architecture: hierarchy when plant architecture is present
    """
    mapping = dict(mapping_json or {})
    column_to_canonical = mapping.get("column_to_canonical") or {}
    timestamp_column = mapping.get("timestamp_column")

    plant_wrap = dict(plant_config_json or {})
    plant = plant_wrap.get("plant") if isinstance(plant_wrap.get("plant"), dict) else plant_wrap
    if not isinstance(plant, dict):
        plant = None

    scada_rows: list[list[Any]] = list(scada_rows_override or [])

    if not scada_rows and raw_csv is not None and raw_csv.exists():
        raw_headers = read_raw_csv_headers(raw_csv)
        source_for_official = resolve_source_columns_for_official(
            column_to_canonical, timestamp_column, raw_headers
        )
        if any(source_for_official.values()):
            scada_rows = remap_csv_to_scada_rows(raw_csv, source_for_official, max_rows)

    return build_parsed_workbook_bytes(
        scada_rows=scada_rows,
        column_to_canonical=column_to_canonical,
        timestamp_column=timestamp_column,
        plant=plant,
    )
