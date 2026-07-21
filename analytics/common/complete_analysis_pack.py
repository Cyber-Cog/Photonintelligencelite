"""Complete Analysis Pack — official SCADA template for running every PIC Lite fault.

Downloadable as multi-sheet Excel and as a ZIP of CSVs. Sample rows use the same
long-format conventions as the demo generator so mapping auto-scores cleanly.
"""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import datetime, timedelta
from typing import Iterable

from openpyxl import Workbook

from analytics.common.architecture_excel import TEMPLATE_COLUMNS

# ---------------------------------------------------------------------------
# Official column headers (human-readable; aliases.yaml maps them → canonical)
# ---------------------------------------------------------------------------

SCADA_COLUMNS: tuple[str, ...] = (
    "Timestamp",
    "Equipment ID",
    "AC Power (kW)",
    "DC Power (kW)",
    "DC Current (A)",
    "DC Voltage (V)",
    "Irradiance (W/m2)",
    "GHI (W/m2)",
    "Module Temp (C)",
    "Ambient Temp (C)",
)

# Official header → canonical field. Single source for demo mapping, tests, and
# scripts/check_template_sync.py — keep in lockstep with docs/COMPLETE_DATA_FORMAT.md.
OFFICIAL_COLUMN_TO_CANONICAL: dict[str, str] = {
    "Timestamp": "timestamp",
    "Equipment ID": "device_id",
    "AC Power (kW)": "ac_power_kw",
    "DC Power (kW)": "dc_power_kw",
    "DC Current (A)": "dc_current_a",
    "DC Voltage (V)": "dc_voltage_v",
    "Irradiance (W/m2)": "poa_w_m2",
    "GHI (W/m2)": "ghi_w_m2",
    "Module Temp (C)": "module_temp_c",
    "Ambient Temp (C)": "ambient_temp_c",
}

assert tuple(OFFICIAL_COLUMN_TO_CANONICAL.keys()) == SCADA_COLUMNS

# Core headers that strongly indicate the official pack (not a generic OEM export).
_PACK_CORE_COLUMNS: tuple[str, ...] = (
    "Timestamp",
    "Equipment ID",
    "AC Power (kW)",
    "DC Current (A)",
)

EXCEL_FILENAME = "pic_lite_complete_analysis_pack.xlsx"
ZIP_FILENAME = "pic_lite_complete_analysis_pack.zip"
SCADA_CSV_NAME = "01_scada_long.csv"
ARCHITECTURE_XLSX_NAME = "02_architecture.xlsx"
README_NAME = "README.txt"

# Compact smoke plant: 2 INV × 2 SCB × 2 STR, one sunny morning (15-min steps).
_SAMPLE_INVERTERS = ("INV-01", "INV-02")
_SCBS_PER_INV = 2
_STRINGS_PER_SCB = 2
_INVERTER_RATED_KW = 90.0
_STRING_ISC_A = 10.0
_SCB_VOLTAGE_V = 620.0
_SAMPLE_START = datetime(2026, 3, 15, 8, 0, 0)
_SAMPLE_STEPS = 8  # 08:00 → 09:45
_SAMPLE_INTERVAL_MIN = 15


def _sample_timestamps() -> list[datetime]:
    return [
        _SAMPLE_START + timedelta(minutes=_SAMPLE_INTERVAL_MIN * i) for i in range(_SAMPLE_STEPS)
    ]


def _irradiance_for(ts: datetime) -> float:
    """Simple rising morning POA profile (W/m²)."""
    hour = ts.hour + ts.minute / 60.0
    # 08:00 ≈ 350, 10:00 ≈ 700
    return round(max(0.0, (hour - 6.0) * 175.0), 1)


def iter_scada_sample_rows() -> Iterable[list]:
    """Yield SCADA rows (header not included) covering inverter / SCB / string / WMS."""
    for ts in _sample_timestamps():
        poa = _irradiance_for(ts)
        ghi = round(poa * 0.97, 1)
        ambient = round(18.0 + poa * 0.012, 1)
        module_t = round(ambient + poa * 0.025, 1)
        frac = poa / 1000.0
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")

        # Plant WMS row (irradiance + temps only)
        yield [
            ts_str,
            "PLANT-WMS-01",
            "",
            "",
            "",
            "",
            poa,
            ghi,
            module_t,
            ambient,
        ]

        for inv in _SAMPLE_INVERTERS:
            inv_dc_kw = 0.0
            for scb_i in range(1, _SCBS_PER_INV + 1):
                scb_id = f"{inv}-SCB-{scb_i:02d}"
                scb_current = 0.0
                for str_i in range(1, _STRINGS_PER_SCB + 1):
                    string_id = f"{scb_id}-STR-{str_i:02d}"
                    i_a = round(frac * _STRING_ISC_A, 3)
                    scb_current += i_a
                    yield [
                        ts_str,
                        string_id,
                        "",
                        round(i_a * _SCB_VOLTAGE_V / 1000.0, 4),
                        i_a,
                        _SCB_VOLTAGE_V,
                        "",
                        "",
                        "",
                        "",
                    ]
                scb_current = round(scb_current, 3)
                scb_p = round(scb_current * _SCB_VOLTAGE_V / 1000.0, 3)
                inv_dc_kw += scb_p
                yield [
                    ts_str,
                    scb_id,
                    "",
                    scb_p,
                    scb_current,
                    _SCB_VOLTAGE_V,
                    "",
                    "",
                    "",
                    "",
                ]
            inv_dc_kw = round(inv_dc_kw, 3)
            ac_kw = round(min(inv_dc_kw * 0.98, _INVERTER_RATED_KW), 3)
            yield [
                ts_str,
                inv,
                ac_kw,
                inv_dc_kw,
                "",
                "",
                "",
                "",
                "",
                "",
            ]


def architecture_sample_rows() -> list[list]:
    rows: list[list] = []
    for inv in _SAMPLE_INVERTERS:
        for scb_i in range(1, _SCBS_PER_INV + 1):
            scb_id = f"{inv}-SCB-{scb_i:02d}"
            note = "Example — replace with your plant IDs" if inv == "INV-01" and scb_i == 1 else ""
            rows.append([inv, _INVERTER_RATED_KW, scb_id, _STRINGS_PER_SCB, note])
    return rows


def _write_text_cell(ws, row: int, column: int, value: str) -> None:
    """Write a literal string cell (never treat leading '=' as a formula).

    openpyxl sets data_type='f' for values starting with '=', which makes Excel
    show the 'We found a problem with some content' repair dialog for README
    separator lines like '=' * 60.
    """
    cell = ws.cell(row=row, column=column, value=value)
    cell.data_type = "s"


def _readme_lines() -> list[str]:
    # Separators must NOT start with '=' — openpyxl stores those as formulas and
    # Excel then fails to open the workbook without repair.
    return [
        "PIC Lite — Complete Analysis Pack (official format)",
        "-" * 60,
        "",
        "WHAT MUST BE PRESENT FOR ALL FAULTS TO RUN",
        "-" * 40,
        "1. Timestamp on every SCADA row (plant local time or UTC — declare timezone in Setup).",
        "2. Equipment ID hierarchy: INV-xx / INV-xx-SCB-xx / INV-xx-SCB-xx-STR-xx (or map columns).",
        "3. Signals (units matter):",
        "   • AC Power (kW)          → KPIs, clipping-by-power, efficiency, box plot",
        "   • DC Power (kW) OR DC Current (A) → efficiency / box plot",
        "   • DC Current (A)         → clipping-by-current, disconnected strings, string outlier",
        "   • DC Voltage (V)         → module damage",
        "   • POA or GHI (W/m²)      → clipping (power & current), disconnected strings",
        "4. Plant architecture in Setup (inverter → SCB → strings_per_scb) + inverter ratings (kW).",
        "   Use 02_architecture.xlsx (or Setup → Download architecture template).",
        "",
        "PER-FAULT CHECKLIST",
        "-" * 40,
        "KPIs                    : AC Power (kW)",
        "Clipping by power       : AC Power + POA/GHI + inverter ratings",
        "Efficiency / box plot   : AC Power + (DC Power OR DC Current)",
        "Clipping by current     : DC Current + POA/GHI + architecture",
        "Disconnected strings    : DC Current + POA/GHI + architecture",
        "Module damage           : DC Voltage + architecture",
        "String outlier          : DC Current",
        "",
        "HOW TO UPLOAD",
        "-" * 40,
        "Accurate path (this pack):",
        "  A) Excel: upload pic_lite_complete_analysis_pack.xlsx — PIC Lite reads the 'scada' sheet.",
        "     Then on Setup → Equipment structure, upload the 'architecture' sheet",
        "     (save that sheet alone as .xlsx, or use 02_architecture.xlsx from the ZIP).",
        "  B) ZIP / CSV: upload 01_scada_long.csv (and optionally other OEM CSVs).",
        "     Upload 02_architecture.xlsx in Setup → Method A.",
        "",
        "Flexible path (messy real SCADA): keep uploading multiple OEM files;",
        "fuzzy column mapping still works. This pack is the official clear path.",
        "",
        "UNITS — always use these (not MW / kA):",
        "  Power kW · Current A · Voltage V · Irradiance W/m² · Temperature °C",
        "",
        "See docs/COMPLETE_DATA_FORMAT.md in the repo for the full specification.",
    ]


def build_scada_csv_text() -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(SCADA_COLUMNS)
    for row in iter_scada_sample_rows():
        writer.writerow(row)
    return buf.getvalue()


def build_excel_bytes() -> bytes:
    """Multi-sheet workbook: README + scada + architecture + fault_checklist."""
    wb = Workbook()

    # --- scada (active first so upload ranking prefers it; README inserted at 0) ---
    ws_scada = wb.active
    ws_scada.title = "scada"
    ws_scada.append(list(SCADA_COLUMNS))
    for row in iter_scada_sample_rows():
        ws_scada.append(row)
    for col in ws_scada.columns:
        ws_scada.column_dimensions[col[0].column_letter].width = 18

    # --- architecture ---
    ws_arch = wb.create_sheet("architecture")
    ws_arch.append(list(TEMPLATE_COLUMNS))
    for row in architecture_sample_rows():
        ws_arch.append(row)
    for col in ws_arch.columns:
        ws_arch.column_dimensions[col[0].column_letter].width = 22

    # --- fault_checklist ---
    ws_chk = wb.create_sheet("fault_checklist")
    ws_chk.append(["Fault / module", "Required SCADA columns", "Required Setup config"])
    checklist = [
        ("Plant KPIs", "AC Power (kW)", "—"),
        ("Clipping by power", "AC Power (kW) + Irradiance/POA or GHI (W/m2)", "Inverter ratings (kW)"),
        ("Inverter efficiency", "AC Power (kW) + DC Power (kW) OR DC Current (A)", "—"),
        ("Box plot", "AC Power (kW) + DC Power (kW) OR DC Current (A)", "—"),
        ("Clipping by current", "DC Current (A) + Irradiance/POA or GHI", "Architecture (INV→SCB→strings)"),
        ("Disconnected strings", "DC Current (A) + Irradiance/POA or GHI", "Architecture"),
        ("Module damage", "DC Voltage (V)", "Architecture"),
        ("String outlier", "DC Current (A)", "—"),
    ]
    for row in checklist:
        ws_chk.append(list(row))
    ws_chk.column_dimensions["A"].width = 28
    ws_chk.column_dimensions["B"].width = 55
    ws_chk.column_dimensions["C"].width = 36

    # --- README first ---
    ws_readme = wb.create_sheet("README", 0)
    for i, line in enumerate(_readme_lines(), start=1):
        _write_text_cell(ws_readme, i, 1, line)
    ws_readme.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_zip_bytes() -> bytes:
    """ZIP: README + long SCADA CSV + architecture Excel (same sample plant)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(README_NAME, "\n".join(_readme_lines()) + "\n")
        zf.writestr(SCADA_CSV_NAME, build_scada_csv_text())
        # Compact architecture matching sample plant (not the large default template)
        arch_wb = Workbook()
        ws = arch_wb.active
        ws.title = "architecture"
        ws.append(list(TEMPLATE_COLUMNS))
        for row in architecture_sample_rows():
            ws.append(row)
        instr = arch_wb.create_sheet("instructions", 0)
        instr["A1"] = "PIC Lite — Plant architecture (Complete Analysis Pack)"
        instr["A3"] = "Upload this file in Setup → Equipment structure → Method A."
        instr["A4"] = "IDs must match Equipment ID values in 01_scada_long.csv."
        instr.column_dimensions["A"].width = 90
        arch_buf = io.BytesIO()
        arch_wb.save(arch_buf)
        zf.writestr(ARCHITECTURE_XLSX_NAME, arch_buf.getvalue())
    return buf.getvalue()


def prefer_scada_sheet_bonus(sheet_name: str) -> float:
    """Boost used by Excel sheet ranking when uploading the official pack."""
    n = (sheet_name or "").strip().lower()
    if n == "scada":
        return 100.0
    if n in {"readme", "instructions", "fault_checklist", "architecture"}:
        return -50.0
    return 0.0


def _normalize_header_key(name: str) -> str:
    s = (name or "").strip().lower()
    s = s.replace("_", " ").replace("-", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def pack_header_overlap(columns: list[str]) -> tuple[int, int]:
    """Return (matched_official_headers, total_official_headers)."""
    detected = {_normalize_header_key(c) for c in columns if (c or "").strip()}
    official = {_normalize_header_key(c) for c in SCADA_COLUMNS}
    return len(detected & official), len(official)


def looks_like_complete_pack(columns: list[str]) -> bool:
    """True when headers look like the official Complete Analysis Pack SCADA sheet.

    Path choice on the Upload page is only a recommendation — callers must use this
    (or equivalent) after upload rather than assuming the pack layout.
    """
    if not columns:
        return False
    matched, total = pack_header_overlap(columns)
    if matched >= max(7, int(total * 0.7)):
        return True
    detected = {_normalize_header_key(c) for c in columns if (c or "").strip()}
    core = {_normalize_header_key(c) for c in _PACK_CORE_COLUMNS}
    return core.issubset(detected) and matched >= 5
