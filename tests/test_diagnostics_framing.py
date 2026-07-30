"""Results diagnostics framing: box plot is analysis, not a fault; Needs data is explicit."""
from __future__ import annotations

from analytics.algorithms import box_plot
from analytics.common.module_kinds import is_analysis_module, is_fault_module, module_kind
from analytics.common.prerequisites import actionable_unavailable_message, evaluate_prerequisites
from analytics.core.result import ResultObject, ResultStatus, ResultTable
from analytics.reports.owner_brief import _has_real_findings, build_owner_brief


def test_box_plot_is_analysis_not_fault():
    assert module_kind("box_plot") == "analysis"
    assert is_analysis_module("box_plot")
    assert not is_fault_module("box_plot")
    assert is_fault_module("disconnected_strings")


def test_box_plot_result_not_owner_finding(demo_context):
    result = box_plot.run(demo_context)
    assert result.status == ResultStatus.OK
    assert result.severity == "info"
    assert result.affected_equipment == []
    assert result.charts and result.charts[0].chart_type == "box"
    # All inverters in one chart / table — not one-inverter-as-fault
    names = {row[0] for row in result.tables[0].rows}
    assert len(names) >= 2
    assert not _has_real_findings(result)

    brief = build_owner_brief([result])
    assert all(p.algorithm_id != "box_plot" for p in brief.problems)


def test_box_plot_stats_table_would_not_count_as_fault_even_with_rows():
    """Regression: stats rows must not become owner findings via severity/equipment."""
    fake = ResultObject(
        algorithm_id="box_plot",
        algorithm_version="1.1.0-port",
        status=ResultStatus.OK,
        title="Box Plot Analysis",
        summary="ok",
        severity="info",
        module_kind="analysis",
        tables=[
            ResultTable(
                title="Efficiency distribution per inverter",
                columns=["Inverter", "Min", "Q1", "Median", "Q3", "Max"],
                rows=[["INV-01", 90, 92, 94, 96, 98], ["INV-02", 88, 90, 91, 93, 95]],
            )
        ],
    )
    assert not _has_real_findings(fake)


def test_needs_data_message_lists_missing_fields():
    msg = actionable_unavailable_message("disconnected_strings", {"dc_current_a", "poa_w_m2"})
    assert msg.startswith("Needs:")
    assert "DC current" in msg
    assert "POA" in msg or "irradiance" in msg.lower()

    rows = evaluate_prerequisites(
        available_fields={"ac_power_kw"},
        has_architecture=False,
        algorithm_ids=["disconnected_strings", "box_plot"],
    )
    by_id = {r["algorithm_id"]: r for r in rows}
    assert by_id["disconnected_strings"]["will_run"] is False
    assert by_id["disconnected_strings"]["message"].startswith("Needs:")
    assert by_id["box_plot"]["module_kind"] == "analysis"
    assert by_id["box_plot"]["will_run"] is False
    assert "Needs:" in by_id["box_plot"]["message"]


def test_pack_checklist_labels_box_plot_as_diagnostic():
    from analytics.common.complete_analysis_pack import build_excel_bytes
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(build_excel_bytes()))
    assert "fault_checklist" in wb.sheetnames
    rows = list(wb["fault_checklist"].iter_rows(values_only=True))
    box_row = next(r for r in rows if r[0] and "box plot" in str(r[0]).lower())
    assert "diagnostic" in str(box_row[0]).lower() or "not a fault" in str(box_row[0]).lower()
