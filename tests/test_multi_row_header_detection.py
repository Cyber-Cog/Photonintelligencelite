"""Multi-row / merged-cell header stitch + string-current channel melt."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from analytics.common.aliasing import score_column
from analytics.common.equipment_ids import derive_level, extract_parent_scb
from analytics.core.context import ResolvedMapping
from analytics.preprocessing.standardize import standardize
from backend.app.services.excel_parser.channels import (
    classify_stitched_column,
    equipment_id_for_channel,
    match_channel_label,
)
from backend.app.services.excel_parser.multi_header import (
    detect_header_block,
    propagate_merged_values,
    stitch_header_rows,
)
from backend.app.services.excel_parser.orchestrator import _run_strategies, convert_excel_to_csv
from backend.app.services.excel_parser.probe import load_sheet_matrix
from backend.app.services.mapping_service import suggest_mapping


# ---------------------------------------------------------------------------
# Channel regex library
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "label,idx",
    [
        ("I1", 1),
        ("I12", 12),
        ("i 24", 24),
        ("STR-03", 3),
        ("String 7", 7),
        ("CH09", 9),
        ("MPPT 2", 2),
        ("SCB_04", 4),
    ],
)
def test_match_channel_label(label: str, idx: int):
    m = match_channel_label(label)
    assert m is not None
    assert m.channel_index == idx


def test_classify_string_current_under_group():
    m = classify_stitched_column(
        group="Strings Current (A)",
        leaf="I5",
        sibling_leaves=[f"I{i}" for i in range(1, 25)],
    )
    assert m is not None
    assert m.field_type == "string_current_channel"
    assert m.channel_index == 5


def test_equipment_id_for_channel_from_sheet_parent():
    assert equipment_id_for_channel("SMB_1", 3) == "SMB-01-STR-03"
    assert derive_level("SMB-01-STR-03") == "string"
    assert extract_parent_scb("SMB-01-STR-03") == "SMB-01"


# ---------------------------------------------------------------------------
# Synthetic openpyxl workbooks
# ---------------------------------------------------------------------------

def _save_merged_two_row_smb(path: Path, *, n_strings: int = 24) -> None:
    """SMB-style: row1 group merges + row2 Voltage/Current/I1..In."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SMB_1"
    # Row 1 groups
    ws["A1"] = "No."
    ws["B1"] = "Date & Time"
    ws["C1"] = "SMB Parameter"
    ws["H1"] = "Strings Current (A)"
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=7)
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=7 + n_strings)
    # Row 2 leaves
    leaves = ["Voltage (V)", "Current (A)", "Power (KW)", "Int. Temp. (C)", "Ext. Temp (C)"]
    for i, leaf in enumerate(leaves):
        ws.cell(2, 3 + i, leaf)
    for i in range(1, n_strings + 1):
        ws.cell(2, 7 + i, f"I{i}")
    # Data
    ws.append([1, "2026-03-15 10:00:00", 650, 40.0, 26.0, 45, 40] + [0.0 if i == 3 else 8.5 for i in range(1, n_strings + 1)])
    ws.append([2, "2026-03-15 11:00:00", 655, 41.0, 27.0, 46, 41] + [0.0 if i == 3 else 9.0 for i in range(1, n_strings + 1)])
    # Persist long enough for string DS path if wired in other tests
    for h in range(12, 14):
        ws.append(
            [h, f"2026-03-15 {h}:00:00", 660, 42.0, 28.0, 47, 42]
            + [0.0 if i == 3 else 9.2 for i in range(1, n_strings + 1)]
        )
    wb.save(path)


def test_two_row_merged_header_stitches_and_detects_24_channels(tmp_path: Path):
    xlsx = tmp_path / "smb_merged.xlsx"
    _save_merged_two_row_smb(xlsx, n_strings=24)

    name, matrix = load_sheet_matrix(xlsx)
    assert name == "SMB_1"
    # After merge propagate, row0 should have Strings Current repeated under I columns
    assert "Strings Current" in " ".join(matrix[0])
    assert matrix[1][7] == "I1"
    assert matrix[1][30] == "I24"

    block = detect_header_block(matrix)
    assert block is not None
    assert block.n_header_rows == 2
    channels = [c for c in block.columns if c.field_type == "string_current_channel"]
    assert len(channels) == 24
    assert channels[0].channel_index == 1
    assert channels[-1].channel_index == 24
    # Primary mapping candidate is the leaf alone
    assert channels[0].primary_candidate.upper() == "I1"

    result = _run_strategies(matrix, sheet_name=name)
    assert result is not None
    assert result.report.strategy == "wide_channel_melt"
    assert result.report.multi_row_header is True
    header = result.rows[0]
    assert "Timestamp" in header
    assert "Equipment ID" in header
    assert "DC Current (A)" in header
    # 4 timestamps × 24 strings
    assert len(result.rows) - 1 == 4 * 24
    eids = {r[header.index("Equipment ID")] for r in result.rows[1:]}
    assert "SMB-01-STR-01" in eids
    assert "SMB-01-STR-24" in eids
    # Disconnected-looking string I3 → STR-03 near zero
    str3 = [r for r in result.rows[1:] if r[header.index("Equipment ID")] == "SMB-01-STR-03"]
    assert all(float(r[header.index("DC Current (A)")]) == 0.0 for r in str3)


def test_one_row_header_regression(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SCADA"
    ws.append(["Timestamp", "Equipment ID", "DC Current (A)", "Irradiance (W/m2)"])
    ws.append(["2026-03-15 10:00:00", "SMB-01", 40.0, 800])
    ws.append(["2026-03-15 10:15:00", "SMB-01", 41.0, 810])
    xlsx = tmp_path / "single.xlsx"
    wb.save(xlsx)

    name, matrix = load_sheet_matrix(xlsx)
    block = detect_header_block(matrix)
    assert block is not None
    assert block.n_header_rows == 1

    result = _run_strategies(matrix, sheet_name=name)
    assert result is not None
    # Must not invent a channel melt
    assert result.report.strategy != "wide_channel_melt"
    assert "DC Current (A)" in result.rows[0]
    assert "Equipment ID" in result.rows[0]


def test_three_row_header_stretch():
    rows = [
        ["Plant Export", "", "", "", "", ""],
        ["", "DC Group", "", "", "Strings Current (A)", ""],
        ["Timestamp", "Voltage (V)", "Current (A)", "Power (kW)", "I1", "I2"],
        ["2026-03-15 10:00:00", "650", "20", "13", "9.1", "0.0"],
        ["2026-03-15 11:00:00", "652", "21", "14", "9.2", "0.0"],
    ]
    # Propagate a synthetic merge on the group row
    merges = [(1, 1, 1, 3), (1, 4, 1, 5)]  # 0-based
    matrix = propagate_merged_values(rows, merges)
    block = detect_header_block(matrix)
    assert block is not None
    assert block.n_header_rows >= 2
    channels = [c for c in block.columns if c.field_type == "string_current_channel"]
    assert len(channels) >= 2

    result = _run_strategies(matrix, sheet_name="SMB_2")
    assert result is not None
    assert result.report.strategy == "wide_channel_melt"
    header = result.rows[0]
    eids = {r[header.index("Equipment ID")] for r in result.rows[1:]}
    assert any("STR-01" in e for e in eids)
    assert any("STR-02" in e for e in eids)


def test_stitch_keeps_leaf_as_primary():
    cols = stitch_header_rows(
        [
            ["Date & Time", "SMB Parameter", "", "Strings Current (A)", ""],
            ["", "Voltage (V)", "Current (A)", "I1", "I2"],
        ]
    )
    by_leaf = {c.leaf_label: c for c in cols}
    assert by_leaf["I1"].primary_candidate == "I1"
    assert by_leaf["I1"].field_type == "string_current_channel"
    assert "Strings Current" in by_leaf["I1"].display_name
    assert by_leaf["Voltage (V)"].primary_candidate == "Voltage (V)"


def test_convert_merged_smb_xlsx_end_to_end(tmp_path: Path):
    xlsx = tmp_path / "smb.xlsx"
    _save_merged_two_row_smb(xlsx, n_strings=8)
    csv_path = tmp_path / "out.csv"
    n = convert_excel_to_csv(xlsx, csv_path, max_decompressed_bytes=5_000_000, max_rows=100_000)
    assert n == 4 * 8
    df = pd.read_csv(csv_path)
    assert "Timestamp" in df.columns
    assert "Equipment ID" in df.columns
    assert "DC Current (A)" in df.columns
    assert df["Equipment ID"].nunique() == 8
    suggestions = {s.column_name: s for s in suggest_mapping(list(df.columns))}
    assert suggestions["DC Current (A)"].canonical_field == "dc_current_a"
    assert suggestions["Equipment ID"].canonical_field == "device_id"


def test_alias_channel_leaf_maps_after_stitch():
    for h in ("I1", "I24", "Strings Current (A) I3", "STR-05"):
        c = score_column(h)
        assert c.canonical_field == "dc_current_a", h


def test_disconnected_string_per_string_zero_when_irradiance_high():
    """I1..In style: string≈0 under POA>50 must confirm disconnected string."""
    from datetime import datetime, timedelta, timezone

    import pandas as pd

    from analytics.algorithms import disconnected_strings
    from analytics.core.result import ResultStatus
    from tests.helpers_fault_context import _empty_rows, build_context

    start = datetime(2026, 3, 15, 10, 0, 0, tzinfo=timezone.utc)
    interval = 15
    n = 8  # 2 hours > 60 min string persistence
    rows: list[dict] = []
    for i in range(n):
        t = start + timedelta(minutes=interval * i)
        for s in range(1, 5):
            sid = f"SMB-01-STR-{s:02d}"
            r = _empty_rows(1)
            r["timestamp_utc"] = [t]
            r["device_id"] = [sid]
            r["device_type"] = ["string"]
            r["inverter_id"] = ["INV-01"]
            r["scb_id"] = ["SMB-01"]
            r["string_id"] = [sid]
            r["dc_current_a"] = [0.0 if s == 2 else 9.0]
            r["dc_voltage_v"] = [620.0]
            rows.append({k: v[0] for k, v in r.items()})
        w = _empty_rows(1)
        w["timestamp_utc"] = [t]
        w["device_id"] = ["WMS-01"]
        w["device_type"] = ["wms"]
        w["poa_w_m2"] = [750.0]
        w["ghi_w_m2"] = [700.0]
        rows.append({k: v[0] for k, v in w.items()})
    frame = pd.DataFrame(rows)
    ctx = build_context(
        frame,
        architecture={"SMB-01": {"inverter_id": "INV-01", "strings_per_scb": 4}},
        sample_interval_minutes=interval,
    )
    result = disconnected_strings.run(ctx)
    assert result.status == ResultStatus.OK
    assert "SMB-01-STR-02" in result.affected_equipment
    assert "string_irradiance_min_w_m2" in result.thresholds_used


@pytest.mark.skipif(
    not Path(r"C:\Users\ayush.r\Downloads\05f3c7a4-3a32-476e-82f7-8d53ca6228af.xlsx").exists(),
    reason="Sample SMB export not present on this machine",
)
def test_real_sample_smb_file_structural_detection():
    sample = Path(r"C:\Users\ayush.r\Downloads\05f3c7a4-3a32-476e-82f7-8d53ca6228af.xlsx")
    name, matrix = load_sheet_matrix(sample)
    assert name == "SMB_1"
    block = detect_header_block(matrix)
    assert block is not None
    assert block.n_header_rows == 2
    channels = [c for c in block.columns if c.field_type == "string_current_channel"]
    assert len(channels) == 24
    result = _run_strategies(matrix, sheet_name=name)
    assert result is not None
    assert result.report.strategy == "wide_channel_melt"
    header = result.rows[0]
    assert "DC Current (A)" in header
    eids = {r[header.index("Equipment ID")] for r in result.rows[1:]}
    assert len([e for e in eids if e.startswith("SMB-01-STR-")]) == 24
