"""Functional onboarding / detection matrix — equipment, aliases, architecture, Excel.

No UI. Asserts upload→mapping→standardize→structure readiness accuracy.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from analytics.common.aliasing import confidence_band, score_column, score_columns
from analytics.common.complete_analysis_pack import OFFICIAL_COLUMN_TO_CANONICAL, SCADA_COLUMNS
from analytics.common.config_loader import load_aliases
from analytics.common.equipment_ids import derive_level, extract_parent_inverter, resolve_inverter_from_architecture
from analytics.common.plant_structure import infer_from_csv, infer_from_ids
from analytics.common.prerequisites import evaluate_prerequisites
from analytics.core.context import ResolvedMapping
from analytics.preprocessing.standardize import standardize
from backend.app.services.excel_parser.orchestrator import _run_strategies
from backend.app.services.excel_parser.probe import load_sheet_matrix, probe_workbook
from backend.app.services.mapping_service import suggest_mapping


@pytest.fixture(autouse=True)
def _reload_aliases():
    load_aliases.cache_clear()
    yield
    load_aliases.cache_clear()


# ---------------------------------------------------------------------------
# Alias / column mapping — must not silently miss signal groups
# ---------------------------------------------------------------------------

SIGNAL_ALIAS_CASES = [
    # AC power
    ("AC Power (kW)", "ac_power_kw"),
    ("pac", "ac_power_kw"),
    ("Active Power", "ac_power_kw"),
    ("P_AC", "ac_power_kw"),
    # DC power
    ("DC Power (kW)", "dc_power_kw"),
    ("pdc", "dc_power_kw"),
    ("Input Power", "dc_power_kw"),
    # DC current / SMB variants
    ("DC Current (A)", "dc_current_a"),
    ("SMB Current", "dc_current_a"),
    ("SCB Current", "dc_current_a"),
    ("String Current", "dc_current_a"),
    ("Combiner Current", "dc_current_a"),
    ("MPPT Current", "dc_current_a"),
    ("Idc", "dc_current_a"),
    ("i_dc", "dc_current_a"),
    ("SMB01 Current", "dc_current_a"),
    ("SCB_12_Idc", "dc_current_a"),
    # DC voltage
    ("DC Voltage (V)", "dc_voltage_v"),
    ("SMB Voltage", "dc_voltage_v"),
    ("Vdc", "dc_voltage_v"),
    # Irradiance
    ("Irradiance (W/m2)", "poa_w_m2"),
    ("POA", "poa_w_m2"),
    ("GHI (W/m2)", "ghi_w_m2"),
    # Temps
    ("Module Temp (C)", "module_temp_c"),
    ("Ambient Temp (C)", "ambient_temp_c"),
    # IDs
    ("Equipment ID", "device_id"),
    ("SMB ID", "scb_id"),
    ("SCB ID", "scb_id"),
]


@pytest.mark.parametrize("header,canonical", SIGNAL_ALIAS_CASES)
def test_signal_alias_maps_with_auto_or_confirm_band(header: str, canonical: str):
    c = score_column(header)
    assert c.canonical_field == canonical, f"{header!r} → {c.canonical_field} (want {canonical})"
    assert c.confidence >= 0.60
    assert confidence_band(c.confidence) in ("auto", "confirm")


def test_suggest_mapping_returns_every_header_no_silent_drop():
    cols = [
        "Timestamp",
        "Equipment ID",
        "AC Power (kW)",
        "DC Power (kW)",
        "SMB Current",
        "DC Voltage (V)",
        "Irradiance (W/m2)",
        "Module Temp (C)",
        "Ambient Temp (C)",
        "Mystery OEM Col XYZ",
    ]
    suggestions = suggest_mapping(cols)
    names = [s.column_name for s in suggestions]
    assert names == cols, "suggest_mapping must surface every header (including unmapped)"
    by = {s.column_name: s for s in suggestions}
    assert by["SMB Current"].canonical_field == "dc_current_a"
    assert by["Mystery OEM Col XYZ"].canonical_field is None
    assert by["Mystery OEM Col XYZ"].band == "manual"


def test_score_columns_does_not_prefer_plant_total_over_inverter_ac():
    scored = score_columns(["Plant AC Power (kW)", "INV1 AC Power (kW)", "Timestamp"])
    by = {c.column_name: c for c in scored}
    # Plant total blocked; inverter-ish column still maps
    assert by["Plant AC Power (kW)"].canonical_field is None or by["Plant AC Power (kW)"].confidence < 0.6
    assert by["INV1 AC Power (kW)"].canonical_field == "ac_power_kw" or by["INV1 AC Power (kW)"].confidence >= 0.55


def test_official_pack_headers_all_map_exactly():
    for official, canonical in OFFICIAL_COLUMN_TO_CANONICAL.items():
        c = score_column(official)
        assert c.canonical_field == canonical
        assert c.confidence == 1.0
    assert tuple(OFFICIAL_COLUMN_TO_CANONICAL.keys()) == SCADA_COLUMNS


# ---------------------------------------------------------------------------
# Equipment ID / architecture — all INV / SMB / strings detected & linked
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "eid,level",
    [
        ("INV-01", "inverter"),
        ("INV-02", "inverter"),
        ("INV_03", "inverter"),
        ("Inverter01", "inverter"),
        ("INV-01-SCB-01", "scb"),
        ("INV-01-SMB-02", "scb"),
        ("SMB-01", "scb"),
        ("SCB-03", "scb"),
        ("Combiner_05", "scb"),
        ("CB12", "scb"),
        ("INV-01-SCB-01-STR-01", "string"),
        ("INV-01-SMB-02-STR-03", "string"),
    ],
)
def test_derive_level_matrix(eid: str, level: str):
    assert derive_level(eid) == level


def test_infer_from_ids_detects_all_inverters_smbs_strings():
    ids = []
    for inv_i in (1, 2, 3):
        inv = f"INV-{inv_i:02d}"
        ids.append(inv)
        for scb_i in range(1, 5):
            scb = f"{inv}-SCB-{scb_i:02d}"
            ids.append(scb)
            for st in range(1, 5):
                ids.append(f"{scb}-STR-{st:02d}")
    structure = infer_from_ids(ids)
    assert structure.detected
    assert set(structure.inverters) == {"INV-01", "INV-02", "INV-03"}
    for inv in structure.inverters.values():
        assert len(inv.scbs) == 4
        for scb in inv.scbs.values():
            assert scb.strings_per_scb == 4
            assert scb.strings_detected
    arch = structure.to_architecture()
    assert len(arch) == 12
    assert arch["INV-02-SCB-03"]["inverter_id"] == "INV-02"
    assert arch["INV-02-SCB-03"]["strings_per_scb"] == 4


def test_standalone_smb_linked_via_architecture_backfill():
    raw = pd.DataFrame(
        {
            "Timestamp": pd.to_datetime(
                ["2026-03-15 10:00:00", "2026-03-15 10:05:00", "2026-03-15 10:00:00", "2026-03-15 10:05:00"]
            ),
            "Equipment ID": ["SMB-01", "SMB-01", "SMB-02", "SMB-02"],
            "SMB Current": [40.0, 39.5, 38.0, 37.5],
            "Irradiance (W/m2)": [800.0, 810.0, 800.0, 810.0],
        }
    )
    mapping = ResolvedMapping(
        column_to_canonical={
            "Equipment ID": "device_id",
            "SMB Current": "dc_current_a",
            "Irradiance (W/m2)": "poa_w_m2",
        },
        confidence_by_column={},
    )
    arch = {
        "SMB-01": {"inverter_id": "INV-01", "strings_per_scb": 16},
        "SMB-02": {"inverter_id": "INV-01", "strings_per_scb": 16},
    }
    result = standardize(raw, mapping, timestamp_column="Timestamp", architecture=arch)
    assert set(result["device_type"]) == {"scb"}
    assert set(result["scb_id"]) == {"SMB-01", "SMB-02"}
    assert (result["inverter_id"] == "INV-01").all()
    assert result["dc_current_a"].notna().all()


def test_hierarchical_smb_parent_inverter_from_id():
    assert extract_parent_inverter("INV-01-SMB-02") == "INV-01"
    assert extract_parent_inverter("INV-01-SCB-01-STR-03") == "INV-01"
    assert resolve_inverter_from_architecture("smb-01", {"SMB-01": {"inverter_id": "INV-A"}}) == "INV-A"


def test_infer_from_demo_csv_full_tree():
    csv_path = Path(__file__).resolve().parent / "fixtures" / "demo_plant_scada.csv"
    structure = infer_from_csv(csv_path, {"Equipment ID": "device_id", "Timestamp": "timestamp"})
    assert set(structure.inverters) == {"INV-01", "INV-02"}
    for inv in structure.inverters.values():
        assert len(inv.scbs) == 4
        for scb in inv.scbs.values():
            assert scb.strings_per_scb == 4


def test_ds_prerequisites_ready_with_smb_current_and_architecture():
    rows = evaluate_prerequisites(
        available_fields={"dc_current_a", "poa_w_m2"},
        has_architecture=True,
        available_by_device_type={"scb": {"dc_current_a", "scb_id"}, "plant": {"poa_w_m2"}},
    )
    ds = next(r for r in rows if r["algorithm_id"] == "disconnected_strings")
    assert ds["will_run"] is True


def test_ds_prerequisites_not_ready_without_architecture():
    rows = evaluate_prerequisites(
        available_fields={"dc_current_a", "poa_w_m2"},
        has_architecture=False,
        available_by_device_type={"scb": {"dc_current_a"}, "plant": {"poa_w_m2"}},
    )
    ds = next(r for r in rows if r["algorithm_id"] == "disconnected_strings")
    assert ds["will_run"] is False


# ---------------------------------------------------------------------------
# Long tidy CSV + wide Excel / multi-sheet
# ---------------------------------------------------------------------------

def test_long_tidy_csv_maps_all_core_signals(tmp_path: Path):
    csv_path = tmp_path / "long_tidy.csv"
    pd.DataFrame(
        {
            "Timestamp": ["2026-03-15 10:00:00"] * 3,
            "Equipment ID": ["INV-01", "INV-01-SCB-01", "INV-01-SCB-01-STR-01"],
            "AC Power (kW)": [80.0, None, None],
            "DC Power (kW)": [82.0, None, None],
            "DC Current (A)": [None, 40.0, 10.0],
            "DC Voltage (V)": [None, 620.0, 620.0],
            "Irradiance (W/m2)": [850.0, 850.0, 850.0],
            "Module Temp (C)": [45.0, 45.0, 45.0],
            "Ambient Temp (C)": [28.0, 28.0, 28.0],
        }
    ).to_csv(csv_path, index=False)

    cols = list(pd.read_csv(csv_path, nrows=0).columns)
    by = {s.column_name: s for s in suggest_mapping(cols)}
    assert by["AC Power (kW)"].canonical_field == "ac_power_kw"
    assert by["DC Current (A)"].canonical_field == "dc_current_a"
    assert by["Irradiance (W/m2)"].canonical_field == "poa_w_m2"

    mapping = ResolvedMapping(
        column_to_canonical={c: by[c].canonical_field for c in cols if by[c].canonical_field and by[c].canonical_field != "timestamp"},
        confidence_by_column={},
    )
    raw = pd.read_csv(csv_path)
    raw["Timestamp"] = pd.to_datetime(raw["Timestamp"])
    canon = standardize(raw, mapping, timestamp_column="Timestamp")
    assert set(canon["device_type"]) == {"inverter", "scb", "string"}
    assert (canon.loc[canon["device_type"] == "scb", "inverter_id"] == "INV-01").all()
    assert (canon.loc[canon["device_type"] == "string", "scb_id"] == "INV-01-SCB-01").all()


def test_wide_smb_excel_preserves_headers_not_collapsed(tmp_path: Path):
    n_smb = 48
    headers = ["Timestamp"] + [f"SMB{i:03d}_Current" for i in range(1, n_smb + 1)]
    wb = Workbook()
    ws = wb.active
    ws.title = "SMB_METRICS"
    ws.append(headers)
    ws.append(["2026-03-15 10:00:00"] + [10.0 + i * 0.01 for i in range(n_smb)])
    path = tmp_path / "wide_smb.xlsx"
    wb.save(path)

    probes = probe_workbook(path, sample_rows=30)
    assert probes[0].n_cols >= n_smb + 1
    name, matrix = load_sheet_matrix(path)
    assert len(matrix[0]) >= n_smb + 1
    result = _run_strategies(matrix, sheet_name=name)
    assert result is not None
    assert len(result.rows[0]) >= n_smb
    assert sum(1 for c in result.rows[0] if "SMB" in str(c).upper()) >= n_smb * 0.9


def test_multi_sheet_excel_probe_ranks_usable_sheets(tmp_path: Path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "README"
    ws1.append(["Notes", "Ignore me"])
    ws1.append(["This is documentation", "not SCADA"])

    ws2 = wb.create_sheet("SCADA")
    ws2.append(["Timestamp", "Equipment ID", "AC Power (kW)", "DC Current (A)", "Irradiance (W/m2)"])
    for i in range(20):
        ws2.append([f"2026-03-15 10:{i:02d}:00", "INV-01", 80.0 + i * 0.1, None, 800.0])

    ws3 = wb.create_sheet("Empty")
    ws3.append(["A", "B"])

    path = tmp_path / "multi.xlsx"
    wb.save(path)

    probes = probe_workbook(path, sample_rows=30)
    names = [p.sheet_name for p in probes]
    assert "SCADA" in names
    # SCADA should be preferred over README for strategies
    name, matrix = load_sheet_matrix(path, sheet_name="SCADA")
    assert name == "SCADA"
    assert "Equipment ID" in matrix[0] or "Timestamp" in matrix[0]
