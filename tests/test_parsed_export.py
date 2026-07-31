"""Tests for parsed Excel export (builder + API endpoint)."""
from __future__ import annotations

import io
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

os.environ["DATABASE_URL"] = "sqlite+pysqlite:///:memory:"
os.environ["SESSION_SECRET"] = "test-secret"
os.environ["AUTH_AUTO_VERIFY"] = "true"
os.environ["PIC_SUPERADMIN_EMAIL"] = ""
os.environ["PIC_SUPERADMIN_PASSWORD"] = ""
os.environ["CORS_ORIGINS"] = "http://testserver"

from analytics.common.complete_analysis_pack import OFFICIAL_COLUMN_TO_CANONICAL, SCADA_COLUMNS
from analytics.common.parsed_export import (
    architecture_hierarchy_rows,
    build_parsed_excel_bytes,
    canonical_frame_to_scada_rows,
    parsed_export_filename,
    source_columns_for_official,
)


def test_parsed_export_filename_short_id():
    assert parsed_export_filename("abcdefgh-1234") == "pic_lite_parsed_abcdefgh.xlsx"


def test_parsed_export_filename_with_plant():
    name = parsed_export_filename("abcdefgh-1234", "Acme Solar / Site 1")
    assert name.startswith("pic_lite_parsed_Acme_Solar")
    assert name.endswith("_abcdefgh.xlsx")


def test_source_columns_for_official_uses_mapping():
    headers = ["DateTime", "Inv", "P_AC", "Noise"]
    mapping = {
        "DateTime": "timestamp",
        "Inv": "device_id",
        "P_AC": "ac_power_kw",
        "Noise": "ignore",
    }
    src = source_columns_for_official(mapping, "DateTime", headers)
    assert src["Timestamp"] == "DateTime"
    assert src["Equipment ID"] == "Inv"
    assert src["AC Power (kW)"] == "P_AC"
    assert src["DC Power (kW)"] is None


def test_source_columns_pack_headers_without_mapping():
    headers = list(SCADA_COLUMNS)
    src = source_columns_for_official({}, None, headers)
    for col in SCADA_COLUMNS:
        assert src[col] == col


def test_build_parsed_excel_bytes_scada_and_companions(tmp_path: Path):
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(
        "DateTime,Inv,P_AC,POA\n"
        "2024-01-01 10:00:00,INV-01,12.5,800\n"
        "2024-01-01 10:15:00,INV-01,13.0,820\n",
        encoding="utf-8",
    )
    mapping = {
        "column_to_canonical": {
            "DateTime": "timestamp",
            "Inv": "device_id",
            "P_AC": "ac_power_kw",
            "POA": "poa_w_m2",
        },
        "timestamp_column": "DateTime",
    }
    plant_config = {
        "plant": {
            "plant_name": "Test Plant",
            "ac_capacity_mw": 0.18,
            "dc_capacity_mwp": 0.216,
            "inverter_capacity_kw": 90.0,
            "equipment_ratings": {"INV-01": 90.0},
            "architecture": {
                "INV-01-SCB-01": {
                    "inverter_id": "INV-01",
                    "strings_per_scb": 2,
                    "dc_capacity_kwp": 54.0,
                },
            },
        }
    }
    raw = build_parsed_excel_bytes(
        raw_csv=csv_path,
        mapping_json=mapping,
        plant_config_json=plant_config,
    )
    wb = load_workbook(io.BytesIO(raw))
    assert "scada" in wb.sheetnames
    assert "column_mapping" in wb.sheetnames
    assert "architecture" in wb.sheetnames

    scada = wb["scada"]
    assert [c.value for c in scada[1]] == list(SCADA_COLUMNS)
    assert scada.max_row == 3
    assert scada["A2"].value == "2024-01-01 10:00:00"
    assert scada["B2"].value == "INV-01"
    assert float(scada["C2"].value) == 12.5
    # Irradiance (W/m2) is column G (index 7)
    assert float(scada["G2"].value) == 800

    mapping_sheet = wb["column_mapping"]
    headers = [c.value for c in mapping_sheet[1]]
    assert headers == ["Source column", "Canonical field", "Official pack header"]
    sources = {mapping_sheet.cell(r, 1).value for r in range(2, mapping_sheet.max_row + 1)}
    assert "DateTime" in sources
    assert "P_AC" in sources

    arch = wb["architecture"]
    assert [c.value for c in arch[1]] == [
        "id",
        "parent_id",
        "device_type",
        "ac_capacity_kw",
        "dc_capacity_kwp",
        "strings_per_scb",
        "notes",
    ]
    assert arch["A2"].value == "PLANT"
    assert arch["C2"].value == "plant"
    assert float(arch["D2"].value) == 180.0  # MW → kW


def test_build_parsed_excel_bytes_oem_headers_without_saved_mapping(tmp_path: Path):
    """Setup download before Continue: infer mapping from CSV headers."""
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(
        "Date And Time,Inverter,Pac_kW,POA\n"
        "2024-01-01 10:00:00,INV-01,12.5,800\n"
        "2024-01-01 10:15:00,INV-01,13.0,820\n",
        encoding="utf-8",
    )
    raw = build_parsed_excel_bytes(raw_csv=csv_path, mapping_json=None)
    wb = load_workbook(io.BytesIO(raw))
    scada = wb["scada"]
    assert scada.max_row == 3
    assert scada["A2"].value == "2024-01-01 10:00:00"
    assert scada["B2"].value == "INV-01"
    assert float(scada["C2"].value) == 12.5


def test_build_parsed_excel_bytes_bom_and_case_insensitive_headers(tmp_path: Path):
    csv_path = tmp_path / "input.csv"
    csv_path.write_bytes(
        "\ufefftimestamp,equipment id,ac power (kw)\n"
        "2024-01-01 10:00:00,INV-01,9.5\n".encode("utf-8")
    )
    raw = build_parsed_excel_bytes(raw_csv=csv_path, mapping_json={})
    wb = load_workbook(io.BytesIO(raw))
    scada = wb["scada"]
    assert scada.max_row == 2
    assert scada["A2"].value == "2024-01-01 10:00:00"
    assert scada["B2"].value == "INV-01"


def test_parsed_xlsx_endpoint_without_saved_mapping(client, tmp_path):
    """Download parsed Excel on Setup before mapping is saved to the job."""
    csrf = _signup(client)
    csv = (
        b"Date And Time,Inverter,Pac_kW,POA\n"
        b"2024-01-01 10:00:00,INV-01,12.5,800\n"
    )
    up = client.post(
        "/api/upload",
        files={"file": ("oem.csv", csv, "text/csv")},
        headers={"X-CSRF-Token": csrf},
    )
    assert up.status_code == 200, up.text
    job_id = up.json()["job_id"]

    resp = client.get(f"/api/jobs/{job_id}/parsed.xlsx")
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    scada = wb["scada"]
    assert scada.max_row >= 2
    assert scada["A2"].value == "2024-01-01 10:00:00"


def test_canonical_frame_to_scada_rows_uses_string_id_for_equipment():
    rows = canonical_frame_to_scada_rows(
        [
            {
                "timestamp_utc": "2024-01-01 10:00:00",
                "string_id": "SMB-01-STR-01",
                "dc_current_a": 8.5,
                "dc_voltage_v": 620.0,
            }
        ]
    )
    assert len(rows) == 1
    assert rows[0][0] == "2024-01-01 10:00:00"
    assert rows[0][1] == "SMB-01-STR-01"
    assert float(rows[0][4]) == 8.5
    assert float(rows[0][5]) == 620.0


def test_export_prefers_canonical_when_raw_remap_is_hollow(tmp_path: Path):
    """Wide/OEM raw CSV with stale mapping should not block parquet export."""
    from backend.app.services.parsed_export_service import export_parsed_excel

    raw_csv = tmp_path / "raw" / "input.csv"
    raw_csv.parent.mkdir(parents=True)
    # Stale pack mapping keys that do not exist in this wide-shaped CSV.
    raw_csv.write_text(
        "FooBar,BazQux\n1,2\n",
        encoding="utf-8",
    )
    canonical_dir = tmp_path / "canonical" / "device_type=string" / "date=2024-01-01"
    canonical_dir.mkdir(parents=True)
    import pandas as pd

    pd.DataFrame(
        [
            {
                "timestamp_utc": "2024-01-01 10:00:00",
                "string_id": "SMB-01-STR-01",
                "dc_current_a": 8.1,
                "dc_voltage_v": 620.0,
            }
        ]
    ).to_parquet(canonical_dir / "part.parquet", index=False)

    mapping = {
        "column_to_canonical": dict(OFFICIAL_COLUMN_TO_CANONICAL),
        "timestamp_column": "Timestamp",
    }
    content, _ = export_parsed_excel(
        job_id="test-job-id",
        raw_csv=raw_csv,
        canonical_dir=tmp_path / "canonical",
        mapping_json=mapping,
        plant_config_json=None,
    )
    wb = load_workbook(io.BytesIO(content))
    scada = wb["scada"]
    assert scada.max_row >= 2
    assert scada["A2"].value == "2024-01-01 10:00:00"
    assert scada["B2"].value == "SMB-01-STR-01"


def test_architecture_hierarchy_rows_empty_without_architecture():
    assert architecture_hierarchy_rows({"plant_name": "X"}) == []


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("JOB_ROOT", str(tmp_path / "uploads"))
    monkeypatch.setenv("DATABASE_URL", "sqlite+pysqlite:///:memory:")
    monkeypatch.setenv("AUTH_AUTO_VERIFY", "true")
    monkeypatch.setenv("PIC_SUPERADMIN_EMAIL", "")

    from backend.app import config, database
    from backend.app.database import Base
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool

    config.get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    TestingSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    database.engine = engine
    database.SessionLocal = TestingSession
    monkeypatch.setattr(database, "_ensure_job_columns", lambda: None)

    from backend.app import models  # noqa: F401

    Base.metadata.create_all(bind=engine)

    from backend.app.main import app
    from backend.app.database import get_db

    def override_db():
        db = TestingSession()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_db
    monkeypatch.setattr("backend.app.main.reclaim_stale_jobs", lambda *_a, **_k: 0)
    monkeypatch.setattr("backend.app.main.seed_superadmin", lambda *_a, **_k: None)

    class _DummyRunner:
        async def start(self):
            return None

        async def stop(self):
            return None

        def submit(self, _job_id: str):
            return None

        def queue_position(self, _job_id: str):
            return None

        def estimated_wait_seconds(self, _job_id: str):
            return None

    monkeypatch.setattr("backend.app.main.get_runner", lambda *_a, **_k: _DummyRunner())
    monkeypatch.setattr(
        "backend.app.main.periodic_cleanup_loop",
        lambda *_a, **_k: __import__("asyncio").sleep(3600),
    )

    from backend.app.auth import rate_limit as rate_limit_mod

    rate_limit_mod._hits.clear()

    from fastapi.testclient import TestClient

    with TestClient(app) as c:
        yield c

    app.dependency_overrides.clear()
    config.get_settings.cache_clear()
    rate_limit_mod._hits.clear()


def _signup(client):
    r = client.post(
        "/api/auth/signup",
        json={"email": "parsed-export@example.com", "password": "password123", "name": "Export"},
    )
    assert r.status_code == 200, r.text
    return r.json()["csrf_token"]


def test_parsed_xlsx_endpoint_round_trip_headers(client, tmp_path):
    csrf = _signup(client)
    # Official pack headers so upload lands tidy without OEM remapping.
    header = ",".join(SCADA_COLUMNS)
    row = "2024-01-01 10:00:00,INV-01,10,11,5,600,900,880,40,25"
    csv = f"{header}\n{row}\n".encode("utf-8")
    up = client.post(
        "/api/upload",
        files={"file": ("pack.csv", csv, "text/csv")},
        headers={"X-CSRF-Token": csrf},
    )
    assert up.status_code == 200, up.text
    job_id = up.json()["job_id"]

    # Persist mapping like Setup Continue would.
    from backend.app.database import SessionLocal
    from backend.app.models import Job

    db = SessionLocal()
    try:
        job = db.get(Job, job_id)
        assert job is not None
        job.mapping_json = {
            "column_to_canonical": dict(OFFICIAL_COLUMN_TO_CANONICAL),
            "confidence_by_column": {},
            "timestamp_column": "Timestamp",
            "detected_oem_signature": None,
        }
        job.plant_config_json = {
            "plant": {
                "plant_name": "Export Plant",
                "ac_capacity_mw": 0.09,
                "dc_capacity_mwp": 0.1,
                "inverter_capacity_kw": 90.0,
                "equipment_ratings": {"INV-01": 90.0},
                "architecture": {
                    "INV-01-SCB-01": {"inverter_id": "INV-01", "strings_per_scb": 2},
                },
            },
            "threshold_overrides": {},
        }
        db.commit()
    finally:
        db.close()

    resp = client.get(f"/api/jobs/{job_id}/parsed.xlsx")
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    cd = resp.headers.get("content-disposition", "")
    assert "pic_lite_parsed_" in cd
    assert ".xlsx" in cd

    wb = load_workbook(io.BytesIO(resp.content))
    assert list(next(wb["scada"].iter_rows(min_row=1, max_row=1, values_only=True))) == list(SCADA_COLUMNS)
    assert wb["scada"].max_row >= 2
    assert "column_mapping" in wb.sheetnames
    assert "architecture" in wb.sheetnames


def test_parsed_xlsx_404_without_data(client):
    csrf = _signup(client)
    # Create a job via upload then delete raw file to simulate missing data.
    csv = b"Timestamp,Equipment ID,AC Power (kW)\n2024-01-01 10:00:00,INV-1,1.0\n"
    up = client.post(
        "/api/upload",
        files={"file": ("t.csv", csv, "text/csv")},
        headers={"X-CSRF-Token": csrf},
    )
    assert up.status_code == 200, up.text
    job_id = up.json()["job_id"]

    from backend.app.config import get_settings
    from backend.app.services.storage import job_paths

    paths = job_paths(get_settings().job_root_path, job_id)
    raw = paths.raw_dir / "input.csv"
    if raw.exists():
        raw.unlink()

    resp = client.get(f"/api/jobs/{job_id}/parsed.xlsx")
    assert resp.status_code == 404
