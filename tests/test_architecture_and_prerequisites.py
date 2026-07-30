"""Architecture Excel template / parse / pattern + prerequisite honesty."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from analytics.common.architecture_excel import (
    HIERARCHY_COLUMNS,
    apply_smb_pattern,
    build_template_bytes,
    parse_architecture_excel,
)
from analytics.common.prerequisites import evaluate_prerequisites, missing_fields_for_algorithm


def test_template_roundtrip():
    raw = build_template_bytes(example_inverters=2, scbs_per_inverter=3, strings_per_scb=8, default_rated_kw=1500)
    parsed = parse_architecture_excel(raw)
    assert parsed.ok
    assert parsed.format == "flat"
    assert len(parsed.inverters) == 2
    assert len(parsed.architecture) == 6
    assert parsed.equipment_ratings["INV-01"] == 1500.0
    assert parsed.architecture["INV-01-SCB-01"]["strings_per_scb"] == 8


def test_hierarchy_sheet_roundtrip():
    wb = Workbook()
    ws = wb.active
    ws.title = "architecture"
    ws.append(list(HIERARCHY_COLUMNS))
    ws.append(["DemoPlant", "", "plant", 180, 216, "", "Demo"])
    ws.append(["INV-01", "DemoPlant", "inverter", 90, 108, "", ""])
    ws.append(["INV-01-SCB-01", "INV-01", "scb", "", 54, 2, ""])
    ws.append(["INV-01-SCB-01-STR-01", "INV-01-SCB-01", "string", "", 27, "", ""])
    ws.append(["INV-01-SCB-01-STR-02", "INV-01-SCB-01", "string", "", 27, "", ""])
    buf = BytesIO()
    wb.save(buf)
    parsed = parse_architecture_excel(buf.getvalue())
    assert parsed.ok
    assert parsed.format == "hierarchy"
    assert parsed.plant_name == "Demo"  # notes preferred over id for display name
    assert parsed.ac_capacity_mw == 0.18
    assert abs(parsed.dc_capacity_mwp - 0.216) < 1e-9
    assert parsed.equipment_ratings["INV-01"] == 90.0
    assert parsed.architecture["INV-01-SCB-01"]["strings_per_scb"] == 2
    draft = parsed.to_plant_config_draft()
    assert draft["architecture_imported"] is True
    assert draft["inverter_capacity_kw"] == 90.0


def test_pattern_applies_to_selected():
    existing = [
        {
            "inverter_id": "KEEP-01",
            "rated_kw": 1000,
            "scbs": [{"scb_id": "KEEP-01-SCB-01", "strings_per_scb": 10, "strings_detected": False}],
        }
    ]
    out = apply_smb_pattern(
        ["INV-01", "INV-02"],
        smbs_per_inverter=2,
        strings_per_smb=24,
        rated_kw=1500,
        existing=existing,
    )
    ids = {i["inverter_id"] for i in out}
    assert ids == {"KEEP-01", "INV-01", "INV-02"}
    inv01 = next(i for i in out if i["inverter_id"] == "INV-01")
    assert len(inv01["scbs"]) == 2
    assert inv01["scbs"][0]["strings_per_scb"] == 24


def test_prerequisites_block_clipping_current_without_irr():
    rows = evaluate_prerequisites(
        available_fields={"dc_current_a"},
        has_architecture=True,
        has_equipment_ratings=True,
        algorithm_ids=["clipping_current"],
    )
    assert len(rows) == 1
    assert rows[0]["will_run"] is False
    assert rows[0]["message"].startswith("Needs:")
    assert "POA" in rows[0]["message"] or "irradiance" in rows[0]["message"].lower() or "poa" in str(rows[0]["missing_fields"]).lower()


def test_prerequisites_run_when_complete():
    rows = evaluate_prerequisites(
        available_fields={"dc_current_a", "poa_w_m2", "ac_power_kw", "dc_power_kw"},
        has_architecture=True,
        has_equipment_ratings=True,
        algorithm_ids=["clipping_current", "box_plot"],
    )
    by_id = {r["algorithm_id"]: r for r in rows}
    assert by_id["clipping_current"]["will_run"] is True
    assert by_id["box_plot"]["will_run"] is True


def test_missing_fields_for_orchestrator():
    missing = missing_fields_for_algorithm("clipping_power", {"ac_power_kw"})
    assert "poa_w_m2" in missing or "ghi_w_m2" in missing


def test_indian_flat_synonyms_and_plant_master_sheet():
    """OEM headers + companion sheet name (not 'architecture') in same workbook as SCADA."""
    wb = Workbook()
    scada = wb.active
    scada.title = "scada"
    scada.append(["Timestamp", "Equipment ID", "AC Power (kW)", "Inverter ID", "SCB ID"])
    scada.append(["2024-06-01 10:00:00", "INV-01-SCB-01", 80.0, "INV-01", "INV-01-SCB-01"])
    plant = wb.create_sheet("Plant Master")
    plant.append(["INV", "SMB", "Rating kW", "No of Strings", "DC Capacity"])
    plant.append(["INV-01", "SMB-01", 1500, 24, 90])
    plant.append(["INV-01", "SMB-02", 1500, 24, 90])
    plant.append(["INV-02", "SMB-03", 1250, 20, 75])
    buf = BytesIO()
    wb.save(buf)
    parsed = parse_architecture_excel(buf.getvalue())
    assert parsed.ok
    assert parsed.format == "flat"
    assert parsed.source_sheet == "Plant Master"
    assert len(parsed.architecture) == 3
    assert parsed.equipment_ratings["INV-01"] == 1500.0
    assert parsed.equipment_ratings["INV-02"] == 1250.0
    assert parsed.architecture["SMB-01"]["strings_per_scb"] == 24
    assert parsed.architecture["SMB-01"]["dc_capacity_kwp"] == 90.0
    draft = parsed.to_plant_config_draft()
    assert draft["architecture_imported"] is True
    assert abs(draft["dc_capacity_mwp"] - 0.255) < 1e-9  # 90+90+75 kWp


def test_flat_string_id_counts_strings_per_scb():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inverter List"
    ws.append(["Inverter ID", "SCB ID", "String ID", "Inverter kW"])
    ws.append(["INV-01", "SCB-01", "STR-01", 1000])
    ws.append(["INV-01", "SCB-01", "STR-02", 1000])
    ws.append(["INV-01", "SCB-01", "STR-03", 1000])
    buf = BytesIO()
    wb.save(buf)
    parsed = parse_architecture_excel(buf.getvalue())
    assert parsed.ok
    assert parsed.architecture["SCB-01"]["strings_per_scb"] == 3
    assert parsed.equipment_ratings["INV-01"] == 1000.0


def test_scada_embedded_inv_scb_when_no_companion_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "scada"
    ws.append(["Timestamp", "Inverter ID", "SCB ID", "AC Power (kW)"])
    ws.append(["2024-01-01 10:00:00", "INV-01", "SCB-A", 50])
    ws.append(["2024-01-01 10:05:00", "INV-01", "SCB-A", 55])
    ws.append(["2024-01-01 10:00:00", "INV-01", "SCB-B", 40])
    buf = BytesIO()
    wb.save(buf)
    parsed = parse_architecture_excel(buf.getvalue())
    assert parsed.ok
    assert parsed.format == "scada_embedded"
    assert set(parsed.architecture) == {"SCB-A", "SCB-B"}
    assert parsed.architecture["SCB-A"]["inverter_id"] == "INV-01"


def test_populated_columns_ignores_all_null_schema():
    """Empty schema columns must not count as available (Will run vs Needs input)."""
    import pandas as pd

    from analytics.core.context import CanonicalDataAccess

    df = pd.DataFrame(
        {
            "timestamp_utc": pd.to_datetime(["2024-01-01T00:00:00Z", "2024-01-01T00:01:00Z"]),
            "ac_power_kw": [100.0, 110.0],
            "poa_w_m2": [800.0, 820.0],
            "dc_power_kw": [None, None],
            "dc_current_a": [None, None],
            "device_type": ["inverter", "inverter"],
        }
    )
    access = CanonicalDataAccess.from_frame(df)
    present = access.populated_columns()
    assert "ac_power_kw" in present and "poa_w_m2" in present
    assert "dc_power_kw" not in present
    assert "dc_current_a" not in present

    rows = evaluate_prerequisites(
        available_fields=present,
        has_architecture=True,
        has_equipment_ratings=True,
    )
    by_id = {r["algorithm_id"]: r for r in rows}
    assert by_id["clipping_power"]["will_run"] is True
    assert by_id["inverter_efficiency"]["will_run"] is False
    assert by_id["disconnected_strings"]["will_run"] is False
