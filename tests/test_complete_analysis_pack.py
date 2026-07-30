"""Complete Analysis Pack template build, mapping auto-score, Excel parse preference."""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from analytics.common.aliasing import score_column
from analytics.common.architecture_excel import (
    HIERARCHY_COLUMNS,
    parse_architecture_excel,
)
from analytics.common.complete_analysis_pack import (
    OFFICIAL_COLUMN_TO_CANONICAL,
    SCADA_COLUMNS,
    build_excel_bytes,
    build_scada_csv_text,
    build_zip_bytes,
    prefer_scada_sheet_bonus,
)
from analytics.common.prerequisites import evaluate_prerequisites
from backend.app.services.excel_parser import parse_excel_to_csv
from backend.app.services.mapping_service import suggest_mapping
from backend.app.services.pack_architecture_import import plant_config_from_architecture_file


def test_excel_pack_sheets_and_sample():
    raw = build_excel_bytes()
    assert raw[:2] == b"PK"
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        assert zf.testzip() is None
        assert "[Content_Types].xml" in zf.namelist()
    wb = load_workbook(io.BytesIO(raw), read_only=False)
    names = set(wb.sheetnames)
    assert {"README", "scada", "architecture", "fault_checklist"} <= names
    ws = wb["scada"]
    rows = list(ws.iter_rows(values_only=True))
    assert list(rows[0]) == list(SCADA_COLUMNS)
    assert len(rows) > 10  # header + sample telemetry
    arch_header = [c.value for c in next(wb["architecture"].iter_rows(max_row=1))]
    assert arch_header == list(HIERARCHY_COLUMNS)
    # README must be literal text — leading '=' becomes formula type and Excel repair-fails
    for row in wb["README"].iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value is None:
            continue
        assert cell.data_type != "f", f"README formula cell: {cell.value!r}"
        assert not (isinstance(cell.value, str) and cell.value.startswith("="))
    wb.close()


def test_pack_architecture_hierarchy_parses_capacities():
    raw = build_excel_bytes()
    parsed = parse_architecture_excel(raw)
    assert parsed.ok
    assert parsed.format == "hierarchy"
    assert "INV-01" in parsed.equipment_ratings
    assert parsed.equipment_ratings["INV-01"] == 90.0
    assert parsed.architecture["INV-01-SCB-01"]["strings_per_scb"] == 2
    assert parsed.architecture["INV-01-SCB-01"].get("dc_capacity_kwp") == 54.0
    # Plant 180 kW / 216 kWp → 0.18 MW / 0.216 MWp
    assert parsed.ac_capacity_mw == 0.18
    assert abs(parsed.dc_capacity_mwp - 0.216) < 1e-9
    assert parsed.inverter_capacity_kw == 90.0

    draft = parsed.to_plant_config_draft()
    assert draft["architecture_imported"] is True
    assert draft["ac_capacity_mw"] == 0.18
    assert draft["equipment_ratings"]["INV-02"] == 90.0


def test_zip_pack_contents():
    raw = build_zip_bytes()
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        names = set(zf.namelist())
        assert "README.txt" in names
        assert "01_scada_long.csv" in names
        assert "02_architecture.xlsx" in names
        csv_text = zf.read("01_scada_long.csv").decode("utf-8")
        assert csv_text.splitlines()[0] == ",".join(SCADA_COLUMNS)
        arch = parse_architecture_excel(zf.read("02_architecture.xlsx"))
        assert arch.ok
        assert arch.format == "hierarchy"
        assert "INV-01" in arch.equipment_ratings
        assert arch.architecture["INV-01-SCB-01"]["strings_per_scb"] == 2
        assert arch.ac_capacity_mw == 0.18


def test_official_headers_auto_map():
    suggestions = suggest_mapping(list(SCADA_COLUMNS))
    by_col = {s.column_name: s for s in suggestions}
    for col, field in OFFICIAL_COLUMN_TO_CANONICAL.items():
        s = by_col[col]
        assert s.canonical_field == field, f"{col} -> {s.canonical_field} (want {field})"
        assert s.band == "auto"
        assert s.confidence >= 0.9


def test_looks_like_complete_pack_detection():
    from analytics.common.complete_analysis_pack import looks_like_complete_pack

    assert looks_like_complete_pack(list(SCADA_COLUMNS)) is True
    assert (
        looks_like_complete_pack(
            ["Timestamp", "Equipment ID", "AC Power (kW)", "DC Current (A)", "GHI (W/m2)"]
        )
        is True
    )
    assert looks_like_complete_pack(["Date And Time", "Inverter", "Pac_kW", "Idc"]) is False


def test_non_pack_fuzzy_stays_confirm_not_silent_auto():
    """OEM-like headers must all appear in suggestions; fuzzy hits are reviewable."""
    cols = ["Date And Time", "INV_Pac_kW", "Some Weird OEM Column", "Status Flag"]
    suggestions = suggest_mapping(cols, pack_match=False)
    assert len(suggestions) == len(cols)
    by_col = {s.column_name: s for s in suggestions}
    weird = by_col["Some Weird OEM Column"]
    assert weird.band == "manual"
    assert weird.canonical_field is None
    for s in suggestions:
        if s.band == "auto":
            assert s.confidence >= 1.0


def test_full_fields_prerequisites_all_will_run():
    rows = evaluate_prerequisites(
        available_fields={
            "ac_power_kw",
            "dc_power_kw",
            "dc_current_a",
            "dc_voltage_v",
            "poa_w_m2",
        },
        has_architecture=True,
        has_equipment_ratings=True,
    )
    assert all(r["will_run"] for r in rows), [r for r in rows if not r["will_run"]]


def test_excel_upload_prefers_scada_sheet(tmp_path: Path):
    assert prefer_scada_sheet_bonus("scada") > prefer_scada_sheet_bonus("architecture")
    excel_path = tmp_path / "pack.xlsx"
    excel_path.write_bytes(build_excel_bytes())
    csv_path = tmp_path / "out.csv"
    n, report = parse_excel_to_csv(
        excel_path,
        csv_path,
        max_decompressed_bytes=50 * 1024 * 1024,
        max_rows=100_000,
    )
    assert n > 10
    assert report.sheet_name == "scada" or "scada" in (report.sheets_probed or [])
    header = csv_path.read_text(encoding="utf-8").splitlines()[0]
    for col in ("Timestamp", "Equipment ID", "AC Power (kW)", "DC Current (A)"):
        assert col in header
    # Architecture still extractable from the same workbook
    draft = plant_config_from_architecture_file(excel_path)
    assert draft is not None
    assert draft["equipment_ratings"]["INV-01"] == 90.0
    assert draft["ac_capacity_mw"] == 0.18


def test_scada_csv_text_nonempty():
    text = build_scada_csv_text()
    assert "INV-01" in text
    assert "PLANT-WMS-01" in text
    # Spot-check alias scoring still exact for official headers
    assert score_column("AC Power (kW)").canonical_field == "ac_power_kw"


def test_single_row_scada_still_maps():
    """Minimal one-row SCADA (no architecture sheet) still maps official headers."""
    suggestions = suggest_mapping(
        ["Timestamp", "Equipment ID", "AC Power (kW)", "Irradiance (W/m2)"]
    )
    by_col = {s.column_name: s.canonical_field for s in suggestions}
    assert by_col["Timestamp"] == "timestamp"
    assert by_col["AC Power (kW)"] == "ac_power_kw"
