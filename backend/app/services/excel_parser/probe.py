"""Probe workbook sheets: dimensions, sample rows, rough dtype guesses."""
from __future__ import annotations

from pathlib import Path

from backend.app.services.excel_parser.headers import cell_to_str, pad_matrix
from backend.app.services.excel_parser.types import SheetProbe


def probe_workbook(excel_path: Path, *, sample_rows: int = 30, max_sheets: int = 8) -> list[SheetProbe]:
    suffix = excel_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _probe_openpyxl(excel_path, sample_rows=sample_rows, max_sheets=max_sheets)
    if suffix == ".xls":
        return _probe_xls(excel_path, sample_rows=sample_rows, max_sheets=max_sheets)
    raise ValueError(f"Unsupported Excel type: {suffix}")


def _probe_openpyxl(excel_path: Path, *, sample_rows: int, max_sheets: int) -> list[SheetProbe]:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        probes: list[SheetProbe] = []
        for name in wb.sheetnames[:max_sheets]:
            ws = wb[name]
            rows: list[list[str]] = []
            n_cols = 0
            # Sample rows only for ranking/dtypes — never truncate column width.
            # openpyxl read_only may omit trailing empty cells per row; track max width
            # across the sample and prefer declared dimensions when available.
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                vals = [cell_to_str(c) for c in row]
                n_cols = max(n_cols, len(vals))
                rows.append(vals)
                if i + 1 >= sample_rows:
                    break
            rows = pad_matrix(rows)
            declared = _declared_max_column(ws)
            n_cols = max(n_cols, declared, max((len(r) for r in rows), default=0))
            if n_cols and rows:
                rows = [list(r) + [""] * (n_cols - len(r)) for r in rows]
            # read_only max_row is often unset until a full pass; sample length is enough to rank.
            n_rows = int(ws.max_row or len(rows) or sample_rows)
            probes.append(
                SheetProbe(
                    sheet_name=name,
                    n_rows=n_rows,
                    n_cols=n_cols,
                    sample_rows=rows,
                    merged_cell_count=0,  # read_only mode has no merged_cells
                    dtype_guesses=_guess_dtypes(rows),
                )
            )
        return probes
    finally:
        wb.close()


def _declared_max_column(ws) -> int:
    """Best-effort sheet width. read_only often leaves max_column unset."""
    try:
        mc = getattr(ws, "max_column", None)
        if mc:
            return int(mc)
    except Exception:  # noqa: BLE001
        pass
    dim = getattr(ws, "dimensions", None)
    if isinstance(dim, str) and ":" in dim:
        # e.g. "A1:YZ500" → parse end column letters
        try:
            end = dim.split(":")[1]
            letters = "".join(ch for ch in end if ch.isalpha())
            if letters:
                return _col_letters_to_index(letters)
        except Exception:  # noqa: BLE001
            return 0
    return 0


def _col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        if not ("A" <= ch <= "Z"):
            break
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _probe_xls(excel_path: Path, *, sample_rows: int, max_sheets: int) -> list[SheetProbe]:
    import pandas as pd

    xl = pd.ExcelFile(excel_path)
    probes: list[SheetProbe] = []
    for name in xl.sheet_names[:max_sheets]:
        df = pd.read_excel(xl, sheet_name=name, header=None, dtype=str, nrows=sample_rows)
        rows = [[cell_to_str(v) for v in row] for row in df.fillna("").to_numpy().tolist()]
        rows = pad_matrix(rows)
        probes.append(
            SheetProbe(
                sheet_name=name,
                n_rows=max(len(rows), 1),
                n_cols=max((len(r) for r in rows), default=0),
                sample_rows=rows,
                merged_cell_count=0,
                dtype_guesses=_guess_dtypes(rows),
            )
        )
    return probes


def _guess_dtypes(rows: list[list[str]]) -> list[str]:
    if len(rows) < 2:
        return []
    width = max(len(r) for r in rows)
    guesses: list[str] = []
    # Dtype guesses sample the first columns only — headers themselves are never truncated.
    for j in range(min(width, 20)):
        vals = [r[j].strip() for r in rows[1:] if j < len(r) and r[j].strip()]
        if not vals:
            guesses.append("empty")
            continue
        numeric = 0
        for v in vals[:20]:
            try:
                float(v.replace(",", ""))
                numeric += 1
            except ValueError:
                pass
        if numeric / max(len(vals[:20]), 1) >= 0.7:
            guesses.append("numeric")
        else:
            guesses.append("text")
    return guesses


def load_sheet_matrix(excel_path: Path, sheet_name: str | None = None) -> tuple[str, list[list[str]]]:
    """Load full sheet as string matrix. Returns (sheet_name, rows).

    Header completeness guarantee: every column that appears in any row is preserved.
    Row sampling is never applied here — only probe uses sample_rows for ranking.
    """
    suffix = excel_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            name = sheet_name or (wb.sheetnames[0] if wb.sheetnames else None)
            if not name:
                raise ValueError("Workbook has no sheets.")
            ws = wb[name]
            declared = _declared_max_column(ws)
            rows: list[list[str]] = []
            max_seen = 0
            # Prefer explicit max_col so trailing header cells are not dropped when
            # early data rows are shorter than the header row.
            iter_kwargs = {}
            if declared > 0:
                iter_kwargs = {"min_col": 1, "max_col": declared}
            for row in ws.iter_rows(values_only=True, **iter_kwargs):
                vals = [cell_to_str(c) for c in row]
                max_seen = max(max_seen, len(vals))
                rows.append(vals)
            width = max(declared, max_seen)
            if width and rows:
                rows = [list(r) + [""] * (width - len(r)) for r in rows]
            else:
                rows = pad_matrix(rows)
            return name, rows
        finally:
            wb.close()

    if suffix == ".xls":
        import pandas as pd

        xl = pd.ExcelFile(excel_path)
        name = sheet_name or xl.sheet_names[0]
        df = pd.read_excel(xl, sheet_name=name, header=None, dtype=str)
        rows = [[cell_to_str(v) for v in row] for row in df.fillna("").to_numpy().tolist()]
        return name, pad_matrix(rows)

    raise ValueError(f"Unsupported Excel type: {suffix}")
