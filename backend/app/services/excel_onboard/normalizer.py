"""Phase 2 — Normalizer: expand merges, ffill, drop blanks, coerce numerics."""
from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from backend.app.services.excel_parser.headers import cell_to_str, pad_matrix
from backend.app.services.excel_parser.multi_header import (
    merged_ranges_from_worksheet,
    propagate_merged_values,
)
from backend.app.services.excel_onboard.analyzer import WorkbookAnalysis

logger = logging.getLogger(__name__)


@dataclass
class NormalizedSheet:
    sheet_name: str
    matrix: list[list[str]]  # string matrix after merge expand + ffill + blank drop
    dropped_blank_cols: list[int]
    dropped_blank_rows: int
    numeric_coerced_cells: int


def normalize_sheet(
    excel_path: Path,
    sheet_name: str,
    *,
    analysis: WorkbookAnalysis | None = None,
) -> NormalizedSheet:
    """Load one sheet, expand merges, ffill, drop empty rows/cols, coerce numeric strings."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        merges = merged_ranges_from_worksheet(ws)
        rows_raw: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows_raw.append(list(row))
    finally:
        wb.close()

    # Expand merges into string matrix
    str_rows = [[cell_to_str(c) for c in r] for r in rows_raw]
    str_rows = pad_matrix(str_rows)
    str_rows = propagate_merged_values(str_rows, merges)

    # Horizontal + vertical forward-fill for sparse header region (first 40 rows)
    header_band = min(40, len(str_rows))
    str_rows = _ffill_horizontal(str_rows, max_row=header_band)
    str_rows = _ffill_vertical_header_band(str_rows, max_row=header_band)

    # Unicode normalize + trim
    for i, row in enumerate(str_rows):
        str_rows[i] = [_norm_text(c) for c in row]

    coerced = _coerce_numeric_strings(str_rows, start_row=_guess_data_start(str_rows, analysis, sheet_name))

    # Drop all-blank rows/cols
    before_rows = len(str_rows)
    str_rows = [r for r in str_rows if any(c.strip() for c in r)]
    dropped_rows = before_rows - len(str_rows)
    str_rows, dropped_cols = _drop_blank_columns(str_rows)

    logger.info(
        "excel_onboard.normalize sheet=%s rows=%s cols=%s merges=%s dropped_rows=%s dropped_cols=%s numeric_coerced=%s",
        sheet_name,
        len(str_rows),
        len(str_rows[0]) if str_rows else 0,
        len(merges),
        dropped_rows,
        len(dropped_cols),
        coerced,
    )
    return NormalizedSheet(
        sheet_name=sheet_name,
        matrix=str_rows,
        dropped_blank_cols=dropped_cols,
        dropped_blank_rows=dropped_rows,
        numeric_coerced_cells=coerced,
    )


def _norm_text(value: str) -> str:
    if not value:
        return ""
    t = unicodedata.normalize("NFKC", value).strip()
    t = re.sub(r"\s+", " ", t)
    return t


def _ffill_horizontal(rows: list[list[str]], *, max_row: int) -> list[list[str]]:
    out: list[list[str]] = []
    for i, row in enumerate(rows):
        if i >= max_row:
            out.append(row)
            continue
        last = ""
        filled: list[str] = []
        for c in row:
            if c.strip():
                last = c
                filled.append(c)
            else:
                filled.append(last)
        out.append(filled)
    return out


def _ffill_vertical_header_band(rows: list[list[str]], *, max_row: int) -> list[list[str]]:
    if not rows:
        return rows
    width = len(rows[0])
    out = [list(r) for r in rows]
    for j in range(width):
        last = ""
        for i in range(min(max_row, len(out))):
            cell = out[i][j] if j < len(out[i]) else ""
            if cell.strip():
                last = cell
            elif last and _looks_headerish(last):
                out[i][j] = last
    return out


def _looks_headerish(value: str) -> bool:
    n = value.lower()
    if any(tok in n for tok in ("icr", "inv", "scb", "smb", "string", "power", "voltage", "current")):
        return True
    return bool(re.match(r"^[A-Za-z]", value)) and not re.match(r"^\d", value)


def _guess_data_start(
    rows: list[list[str]], analysis: WorkbookAnalysis | None, sheet_name: str
) -> int:
    if analysis:
        for s in analysis.sheets:
            if s.sheet_name == sheet_name:
                return max(0, s.first_data_row_estimate)
    return min(10, len(rows))


def _coerce_numeric_strings(rows: list[list[str]], *, start_row: int) -> int:
    """Convert numeric-looking strings in data region to canonical numeric strings."""
    coerced = 0
    for i in range(start_row, len(rows)):
        for j, cell in enumerate(rows[i]):
            if not cell or not cell.strip():
                continue
            s = cell.strip().replace(",", "")
            if re.fullmatch(r"-?\d+(\.\d+)?([eE][+-]?\d+)?", s):
                try:
                    f = float(s)
                    # Keep compact representation
                    new = str(int(f)) if f.is_integer() and abs(f) < 1e15 else str(f)
                    if new != cell:
                        rows[i][j] = new
                        coerced += 1
                except ValueError:
                    pass
    return coerced


def _drop_blank_columns(rows: list[list[str]]) -> tuple[list[list[str]], list[int]]:
    if not rows:
        return rows, []
    width = len(rows[0])
    keep: list[int] = []
    dropped: list[int] = []
    for j in range(width):
        if any((r[j].strip() if j < len(r) else "") for r in rows):
            keep.append(j)
        else:
            dropped.append(j)
    if len(keep) == width:
        return rows, []
    return [[r[j] for j in keep] for r in rows], dropped
