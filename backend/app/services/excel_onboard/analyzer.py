"""Phase 1 — Workbook analyzer (openpyxl, no AI)."""
from __future__ import annotations

import logging
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


@dataclass
class SheetAnalysis:
    sheet_name: str
    n_rows: int
    n_cols: int
    merged_ranges: list[str] = field(default_factory=list)
    header_depth_estimate: int = 1
    first_data_row_estimate: int = 0  # 0-based
    blank_column_indexes: list[int] = field(default_factory=list)
    hidden_row_indexes: list[int] = field(default_factory=list)
    hidden_column_indexes: list[int] = field(default_factory=list)
    freeze_panes: str | None = None


@dataclass
class WorkbookAnalysis:
    path: str
    sheet_names: list[str]
    sheets: list[SheetAnalysis]
    primary_sheet: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "sheet_names": self.sheet_names,
            "primary_sheet": self.primary_sheet,
            "sheets": [asdict(s) for s in self.sheets],
        }


def analyze_workbook(excel_path: Path, *, max_sheets: int = 8) -> WorkbookAnalysis:
    """Inspect workbook structure without sending anything to AI."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    suffix = excel_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Analyzer supports .xlsx/.xlsm only (got {suffix})")

    wb = load_workbook(excel_path, read_only=False, data_only=True)
    try:
        sheets: list[SheetAnalysis] = []
        for name in wb.sheetnames[:max_sheets]:
            ws = wb[name]
            merges = [str(r) for r in (ws.merged_cells.ranges or [])]
            n_rows = int(ws.max_row or 0)
            n_cols = int(ws.max_column or 0)

            # Sample first 40 rows to estimate header depth / first data / blank cols
            sample: list[list[Any]] = []
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(40, n_rows or 40), values_only=True)):
                sample.append(list(row))
            header_depth, first_data = _estimate_header_depth(sample)
            blank_cols = _blank_columns(sample, n_cols)

            hidden_rows = [
                i for i in range(1, min(n_rows, 500) + 1) if ws.row_dimensions[i].hidden
            ]
            hidden_cols: list[int] = []
            for c in range(1, min(n_cols, 300) + 1):
                letter = get_column_letter(c)
                if ws.column_dimensions[letter].hidden:
                    hidden_cols.append(c - 1)

            freeze = None
            if ws.freeze_panes:
                freeze = str(ws.freeze_panes)

            sheets.append(
                SheetAnalysis(
                    sheet_name=name,
                    n_rows=n_rows,
                    n_cols=n_cols,
                    merged_ranges=merges,
                    header_depth_estimate=header_depth,
                    first_data_row_estimate=first_data,
                    blank_column_indexes=blank_cols,
                    hidden_row_indexes=[r - 1 for r in hidden_rows[:200]],
                    hidden_column_indexes=hidden_cols[:200],
                    freeze_panes=freeze,
                )
            )
            logger.info(
                "excel_onboard.analyze sheet=%s rows=%s cols=%s merges=%s header_depth=%s first_data=%s",
                name,
                n_rows,
                n_cols,
                len(merges),
                header_depth,
                first_data,
            )

        primary = _pick_primary(sheets)
        return WorkbookAnalysis(
            path=str(excel_path),
            sheet_names=list(wb.sheetnames),
            sheets=sheets,
            primary_sheet=primary,
        )
    finally:
        wb.close()


def _estimate_header_depth(sample: list[list[Any]]) -> tuple[int, int]:
    """Return (header_depth, first_data_row_0based)."""
    import re
    from datetime import datetime

    if not sample:
        return 1, 0

    inv_re = re.compile(r"^\s*(?:INVERTER|INV)\s*[-_]?\s*\d+\s*$", re.I)
    icr_re = re.compile(r"^\s*ICR\s*[-_]?\s*\d+\s*$", re.I)
    metric_tok = ("power", "voltage", "current", "temp", "irradiance", "energy")

    best_device = -1
    best_leaf = -1
    for i, row in enumerate(sample[:30]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        inv_hits = sum(1 for c in cells if c and inv_re.match(c))
        icr_hits = sum(1 for c in cells if c and icr_re.match(c))
        metric_hits = sum(1 for c in cells if c and any(t in c.lower() for t in metric_tok))
        if inv_hits >= 2:
            best_device = i
        if metric_hits >= 4 and best_device >= 0 and i >= best_device:
            best_leaf = i
            break
        if icr_hits >= 2 and best_device < 0:
            # ICR banner often sits above INV
            pass

    if best_device >= 0 and best_leaf >= best_device:
        depth = best_leaf - max(0, best_device - 1) + 1
        return max(1, depth), best_leaf + 1

    # Fallback: first row that looks like timestamps / numbers
    for i, row in enumerate(sample):
        vals = [c for c in row if c is not None and str(c).strip()]
        if not vals:
            continue
        numericish = 0
        for v in vals[:8]:
            if isinstance(v, (int, float, datetime)):
                numericish += 1
            else:
                s = str(v).strip()
                try:
                    float(s.replace(",", ""))
                    numericish += 1
                except ValueError:
                    pass
        if numericish >= max(2, len(vals[:8]) // 2) and i > 0:
            return max(1, i), i
    return 1, min(1, len(sample))


def _blank_columns(sample: list[list[Any]], n_cols: int) -> list[int]:
    if n_cols <= 0:
        n_cols = max((len(r) for r in sample), default=0)
    blank: list[int] = []
    for j in range(n_cols):
        empty = True
        for row in sample:
            if j < len(row) and row[j] is not None and str(row[j]).strip():
                empty = False
                break
        if empty:
            blank.append(j)
    return blank


def _pick_primary(sheets: list[SheetAnalysis]) -> str | None:
    if not sheets:
        return None
    scored: list[tuple[float, str]] = []
    for s in sheets:
        score = s.n_cols * 0.02 + min(s.n_rows, 5000) * 0.001 + len(s.merged_ranges) * 0.1
        name_l = s.sheet_name.lower()
        if "report" in name_l or "inverter" in name_l or "scada" in name_l:
            score += 5
        scored.append((score, s.sheet_name))
    scored.sort(reverse=True)
    return scored[0][1]
